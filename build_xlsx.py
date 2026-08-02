#!/usr/bin/env python3
import json, datetime, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

HERE = os.path.dirname(os.path.abspath(__file__))
DATASET_PATH = os.path.join(HERE, "dataset.json")
OUT_PATH = os.path.join(HERE, "Investment_Master.xlsx")
DATA = json.load(open(DATASET_PATH))
AS_OF = DATA["meta"]["as_of"]

wb = Workbook()

# ---------- Style helpers ----------
FONT = "Arial"
C_INPUT = Font(name=FONT, color="0000FF")           # blue = hardcoded input
C_FORM  = Font(name=FONT, color="000000")           # black = formula
C_LINK  = Font(name=FONT, color="008000")           # green = link to other sheet
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1F2A44")
SUB_FONT = Font(name=FONT, italic=True, size=9, color="666666")
BOLD = Font(name=FONT, bold=True)
HDR_FILL = PatternFill("solid", fgColor="1F2A44")
SUBHDR_FILL = PatternFill("solid", fgColor="E7ECF3")
YELLOW = PatternFill("solid", fgColor="FFFF99")      # fill-in / live-update cells
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
NEG_FILL = PatternFill("solid", fgColor="FCE4E4")
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

FMT_USD = '$#,##0.00;($#,##0.00);"-"'
FMT_USD0 = '$#,##0;($#,##0);"-"'
FMT_NUM = '#,##0.00;(#,##0.00);"-"'
FMT_NUM0 = '#,##0;(#,##0);"-"'
FMT_PCT = '0.0%'
FMT_PCT2 = '0.00%'
FMT_ARS = '#,##0;(#,##0);"-"'
FMT_DATE = 'yyyy-mm-dd'

def style_header(ws, row, cols, start=1):
    for i, name in enumerate(cols):
        c = ws.cell(row=row, column=start+i, value=name)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CTR; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def per_usd(ccy_ref):
    # units of that currency per 1 USD, from the Rates ccy table
    return f"INDEX(Rates!$F$4:$F$8,MATCH({ccy_ref},Rates!$E$4:$E$8,0))"

def title(ws, text, sub=None, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text); c.font = TITLE_FONT
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        s = ws.cell(row=2, column=1, value=sub); s.font = SUB_FONT

# =====================================================================
# RATES SHEET
# =====================================================================
rs = wb.active
rs.title = "Rates"
title(rs, "FX Rates & Settings", "Single source of truth for all currency conversion. Blue = editable input. Update these to refresh the whole workbook.", span=6)

rs["A4"] = "Currency (per 1 USD)"; rs["A4"].font = BOLD
rs["A5"] = "EUR per USD"; rs["B5"] = DATA["fx"]["EUR_per_USD"]
rs["A6"] = "GBP per USD"; rs["B6"] = DATA["fx"]["GBP_per_USD"]
for r in (5,6):
    rs.cell(row=r, column=2).font = C_INPUT
    rs.cell(row=r, column=2).number_format = '#,##0.00000'
    rs.cell(row=r, column=2).fill = YELLOW

rs["A8"] = "ARS rate options (ARS per USD)"; rs["A8"].font = BOLD
ars = DATA["fx"]["ars_rates"]
ars_order = ["oficial", "mayorista", "mep", "ccl", "blue"]
row = 9
ars_first = row
for k in ars_order:
    rs.cell(row=row, column=1, value=k.upper())
    cell = rs.cell(row=row, column=2, value=ars[k])
    cell.font = C_INPUT; cell.number_format = FMT_ARS; cell.fill = YELLOW
    row += 1
ars_last = row-1

sel_row = row+1
rs.cell(row=sel_row, column=1, value="Selected ARS rate"); rs.cell(row=sel_row,column=1).font = BOLD
sel_cell = rs.cell(row=sel_row, column=2, value=DATA["fx"]["ars_rate_selected"].upper())
sel_cell.font = C_INPUT; sel_cell.fill = YELLOW; sel_cell.alignment = CTR
dv = DataValidation(type="list", formula1='"OFICIAL,MAYORISTA,MEP,CCL,BLUE"', allow_blank=False)
rs.add_data_validation(dv); dv.add(sel_cell)
rs.cell(row=sel_row+1, column=1, value="ARS per USD (used everywhere)")
ars_used = rs.cell(row=sel_row+1, column=2,
    value=f"=INDEX(B{ars_first}:B{ars_last},MATCH(B{sel_row},A{ars_first}:A{ars_last},0))")
ars_used.font = C_FORM; ars_used.number_format = FMT_ARS; ars_used.font = BOLD

disp_row = sel_row+3
rs.cell(row=disp_row, column=1, value="Master display currency"); rs.cell(row=disp_row,column=1).font=BOLD
disp_cell = rs.cell(row=disp_row, column=2, value="USD"); disp_cell.font=C_INPUT; disp_cell.fill=YELLOW; disp_cell.alignment=CTR
dv2 = DataValidation(type="list", formula1='"USD,EUR,ARS"', allow_blank=False)
rs.add_data_validation(dv2); dv2.add(disp_cell)

rs.cell(row=disp_row+2, column=1, value="Last FX update"); rs.cell(row=disp_row+2,column=1).font=BOLD
u = rs.cell(row=disp_row+2, column=2, value=AS_OF); u.font=C_INPUT; u.fill=YELLOW

# ccy -> per USD reference table (headers row 3, data E4:F8)
rs["E3"] = "Ccy"; rs["F3"] = "Per USD"
rs["E3"].font = BOLD; rs["F3"].font = BOLD
ccy_tbl = [("USD", 1), ("EUR", "=B5"), ("ARS", f"=B{sel_row+1}"), ("GBP", "=B6"), ("BRL", 5.12)]
for i,(cc,val) in enumerate(ccy_tbl):
    rs.cell(row=4+i, column=5, value=cc).font = BOLD
    vc = rs.cell(row=4+i, column=6, value=val)
    vc.number_format = '#,##0.00000'
    vc.font = C_LINK if isinstance(val,str) else C_INPUT

# handy named cells (references used by Master)
EUR_REF = "Rates!$B$5"
ARS_REF = f"Rates!$B${sel_row+1}"

rs.column_dimensions["A"].width = 32
rs.column_dimensions["B"].width = 16
rs.column_dimensions["E"].width = 16
rs.column_dimensions["F"].width = 14

# small legend
lg = disp_row+5
rs.cell(row=lg, column=1, value="Legend").font = BOLD
legend = [("Blue text on yellow", "You edit this (prices, rates, positions)"),
          ("Black text", "Formula — do not overwrite"),
          ("Green text", "Links to another sheet")]
for i,(a,b) in enumerate(legend):
    rs.cell(row=lg+1+i, column=1, value=a).font = SUB_FONT
    rs.cell(row=lg+1+i, column=2, value=b).font = SUB_FONT
    rs.merge_cells(start_row=lg+1+i,start_column=2,end_row=lg+1+i,end_column=4)

# store some layout constants to reuse
RATES = dict(EUR="Rates!$B$5", ARS=f"Rates!$B${sel_row+1}")

# =====================================================================
# Generic position-sheet builders
# =====================================================================
def sheet_start(ws, ttl, sub, span):
    title(ws, ttl, sub, span=span)
    return 4  # header row

def totals_row_style(ws, r, cols):
    for c in range(1, cols+1):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = BOLD
        ws.cell(row=r, column=c).border = BORDER

def col_widths(ws, widths):
    for i,w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

# ---------- Equities (Global & ARG share structure) ----------
def build_equities(ws, ttl, sub, rows):
    hr = sheet_start(ws, ttl, sub, span=14)
    cols = ["Ticker","ISIN","Name","Qty","Avg Cost","Price (live)","SoY Price","Ccy",
            "Cost Basis","Market Value","Market Value (USD)","Unreal. P&L (USD)","YTD %","Weight %"]
    style_header(ws, hr, cols)
    first = hr+1
    for i,p in enumerate(rows):
        r = first+i
        ws.cell(row=r,column=1,value=p["ticker"]).font=C_INPUT
        ic=ws.cell(row=r,column=2,value=p.get("isin","")); ic.font=C_INPUT; ic.fill=YELLOW
        ws.cell(row=r,column=3,value=p["name"]).font=C_INPUT
        ws.cell(row=r,column=4,value=p["qty"]).font=C_INPUT
        ws.cell(row=r,column=5,value=p["avg_cost"]).font=C_INPUT
        pc=ws.cell(row=r,column=6,value=p["price"]); pc.font=C_INPUT; pc.fill=YELLOW
        sc=ws.cell(row=r,column=7,value=p.get("price_soy",p["price"])); sc.font=C_INPUT; sc.fill=YELLOW
        ws.cell(row=r,column=8,value=p["ccy"]).font=C_INPUT
        ws.cell(row=r,column=9,value=f"=D{r}*E{r}").font=C_FORM
        ws.cell(row=r,column=10,value=f"=D{r}*F{r}").font=C_FORM
        ws.cell(row=r,column=11,value=f"=J{r}/{per_usd(f'H{r}')}").font=C_FORM
        ws.cell(row=r,column=12,value=f"=(J{r}-I{r})/{per_usd(f'H{r}')}").font=C_FORM
        ws.cell(row=r,column=13,value=f"=IF(G{r}=0,0,(F{r}-G{r})/G{r})").font=C_FORM
        last=r
    tot = last+1
    ws.cell(row=tot,column=1,value="TOTAL")
    ws.cell(row=tot,column=11,value=f"=SUM(K{first}:K{last})")
    ws.cell(row=tot,column=12,value=f"=SUM(L{first}:L{last})")
    for i in range(first,last+1):
        ws.cell(row=i,column=14,value=f"=IF($K${tot}=0,0,K{i}/$K${tot})")
    totals_row_style(ws, tot, 14)
    for r in range(first,last+1):
        ws.cell(row=r,column=4).number_format=FMT_NUM0
        for cc in (5,6,7,9,10): ws.cell(row=r,column=cc).number_format=FMT_NUM
        for cc in (11,12): ws.cell(row=r,column=cc).number_format=FMT_USD
        ws.cell(row=r,column=13).number_format=FMT_PCT
        ws.cell(row=r,column=14).number_format=FMT_PCT
    ws.cell(row=tot,column=11).number_format=FMT_USD
    ws.cell(row=tot,column=12).number_format=FMT_USD
    col_widths(ws,[11,15,24,9,11,12,11,6,13,13,16,16,9,9])
    return f"{ws.title}!K{first}:K{last}", f"'{ws.title}'!$K${tot}", f"'{ws.title}'!$L${tot}"

eqg = wb.create_sheet("EqGlobal")
EQG_RANGE, EQG_TOT, EQG_PNL = build_equities(eqg, "Global Equities",
    "Prices in each instrument's own currency. Yellow price cells are the live-update inputs. YTD = price return since start of year.", DATA["equities_global"])

eqa = wb.create_sheet("EqARG")
EQA_RANGE, EQA_TOT, EQA_PNL = build_equities(eqa, "Argentine Equities (ARS)",
    "Merval / local equities priced in ARS. Converted to USD via selected ARS rate on the Rates tab.", DATA["equities_arg"])

# ---------- Funds (ETFs / mutual / money-market FCI / alternatives) ----------
def build_funds(ws, rows):
    hr=sheet_start(ws,"Funds","ETFs, mutual funds, money-market (FCI) and alternatives. Valued at units (quotas) x NAV. YTD = NAV move since start of year.",span=15)
    cols=["Fund","ISIN","Name","Type","Units","NAV (live)","Cost NAV","SoY NAV","Ccy","Cost Basis","Market Value","MV (USD)","Unreal. P&L (USD)","YTD %","Weight %"]
    style_header(ws,hr,cols); first=hr+1
    for i,p in enumerate(rows):
        r=first+i
        ws.cell(row=r,column=1,value=p["id"]).font=C_INPUT
        ic=ws.cell(row=r,column=2,value=p.get("isin","")); ic.font=C_INPUT; ic.fill=YELLOW
        ws.cell(row=r,column=3,value=p.get("name","")).font=C_INPUT
        ws.cell(row=r,column=4,value=p.get("fund_type","")).font=C_INPUT
        ws.cell(row=r,column=5,value=p["units"]).font=C_INPUT
        pc=ws.cell(row=r,column=6,value=p["nav"]); pc.font=C_INPUT; pc.fill=YELLOW
        ws.cell(row=r,column=7,value=p.get("cost_nav",p["nav"])).font=C_INPUT
        so=ws.cell(row=r,column=8,value=p.get("nav_soy",p["nav"])); so.font=C_INPUT; so.fill=YELLOW
        ws.cell(row=r,column=9,value=p["ccy"]).font=C_INPUT
        ws.cell(row=r,column=10,value=f"=E{r}*G{r}").font=C_FORM
        ws.cell(row=r,column=11,value=f"=E{r}*F{r}").font=C_FORM
        ws.cell(row=r,column=12,value=f"=K{r}/{per_usd(f'I{r}')}").font=C_FORM
        ws.cell(row=r,column=13,value=f"=(K{r}-J{r})/{per_usd(f'I{r}')}").font=C_FORM
        ws.cell(row=r,column=14,value=f"=IF(H{r}=0,0,(F{r}-H{r})/H{r})").font=C_FORM
        last=r
    tot=last+1
    ws.cell(row=tot,column=1,value="TOTAL")
    ws.cell(row=tot,column=12,value=f"=SUM(L{first}:L{last})")
    ws.cell(row=tot,column=13,value=f"=SUM(M{first}:M{last})")
    for i in range(first,last+1):
        ws.cell(row=i,column=15,value=f"=IF($L${tot}=0,0,L{i}/$L${tot})")
    totals_row_style(ws,tot,15)
    for r in range(first,last+1):
        ws.cell(row=r,column=5).number_format=FMT_NUM0
        for cc in (6,7,8,10,11): ws.cell(row=r,column=cc).number_format=FMT_NUM
        for cc in (12,13): ws.cell(row=r,column=cc).number_format=FMT_USD
        ws.cell(row=r,column=14).number_format=FMT_PCT; ws.cell(row=r,column=15).number_format=FMT_PCT
    ws.cell(row=tot,column=12).number_format=FMT_USD; ws.cell(row=tot,column=13).number_format=FMT_USD
    col_widths(ws,[16,15,20,15,10,11,10,10,6,13,14,14,15,9,9])
    return f"'Funds'!$L${tot}", f"'Funds'!$M${tot}"
fns=wb.create_sheet("Funds")
FN_TOT, FN_PNL = build_funds(fns, DATA.get("funds",[]))

# ---------- Bonds (Intl & ARG share structure) ----------
def build_bonds(ws, ttl, sub, rows):
    hr = sheet_start(ws, ttl, sub, span=21)
    cols = ["Bond","ISIN","Name","Face","Coupon %","Freq","Price %","SoY %","Ccy",
            "Last Coupon","Next Coupon","Days to Next","Maturity","YTM","Duration","Mod Dur",
            "Clean Value","Accrued Int.","Dirty (USD)","Accrued (USD)","Price YTD %","Cost %","Unreal. P&L (USD)"]
    style_header(ws, hr, cols)
    first=hr+1
    for i,b in enumerate(rows):
        r=first+i
        ws.cell(row=r,column=1,value=b["id"]).font=C_INPUT
        ic=ws.cell(row=r,column=2,value=b.get("isin","")); ic.font=C_INPUT; ic.fill=YELLOW
        ws.cell(row=r,column=3,value=b["name"]).font=C_INPUT
        ws.cell(row=r,column=4,value=b["face"]).font=C_INPUT
        cp=ws.cell(row=r,column=5,value=b["coupon_pct"]/100); cp.font=C_INPUT
        ws.cell(row=r,column=6,value=b["freq_per_yr"]).font=C_INPUT
        pr=ws.cell(row=r,column=7,value=b["price_pct"]/100); pr.font=C_INPUT; pr.fill=YELLOW
        so=ws.cell(row=r,column=8,value=b.get("price_pct_soy",b["price_pct"])/100); so.font=C_INPUT; so.fill=YELLOW
        ws.cell(row=r,column=9,value=b["ccy"]).font=C_INPUT
        lc=ws.cell(row=r,column=10,value=b["last_coupon"]); lc.font=C_INPUT
        nc=ws.cell(row=r,column=11,value=b["next_coupon"]); nc.font=C_INPUT
        ws.cell(row=r,column=12,value=f'=K{r}-TODAY()').font=C_FORM
        mt=ws.cell(row=r,column=13,value=b["maturity"]); mt.font=C_INPUT
        ws.cell(row=r,column=14,value=f"=IFERROR(YIELD(TODAY(),DATEVALUE(M{r}),E{r},G{r}*100,100,F{r},0),0)").font=C_FORM
        ws.cell(row=r,column=15,value=f"=IFERROR(DURATION(TODAY(),DATEVALUE(M{r}),E{r},N{r},F{r},0),0)").font=C_FORM
        ws.cell(row=r,column=16,value=f"=IFERROR(MDURATION(TODAY(),DATEVALUE(M{r}),E{r},N{r},F{r},0),0)").font=C_FORM
        ws.cell(row=r,column=17,value=f"=D{r}*G{r}").font=C_FORM                       # clean ccy
        ws.cell(row=r,column=18,value=f"=D{r}*E{r}*(TODAY()-J{r})/365").font=C_FORM      # accrued ccy
        ws.cell(row=r,column=19,value=f"=(Q{r}+R{r})/{per_usd(f'I{r}')}").font=C_FORM
        ws.cell(row=r,column=20,value=f"=R{r}/{per_usd(f'I{r}')}").font=C_FORM
        ws.cell(row=r,column=21,value=f"=IF(H{r}=0,0,(G{r}-H{r})/H{r})").font=C_FORM
        vc=ws.cell(row=r,column=22,value=b.get("cost_pct",b["price_pct"])/100); vc.font=C_INPUT; vc.fill=YELLOW
        ws.cell(row=r,column=23,value=f"=D{r}*(G{r}-V{r})/{per_usd(f'I{r}')}").font=C_FORM   # unrealized P&L USD
        last=r
    tot=last+1
    ws.cell(row=tot,column=1,value="TOTAL")
    ws.cell(row=tot,column=16,value=f"=IFERROR(SUMPRODUCT(S{first}:S{last},P{first}:P{last})/SUM(S{first}:S{last}),0)")
    ws.cell(row=tot,column=19,value=f"=SUM(S{first}:S{last})")
    ws.cell(row=tot,column=20,value=f"=SUM(T{first}:T{last})")
    ws.cell(row=tot,column=23,value=f"=SUM(W{first}:W{last})")
    totals_row_style(ws,tot,23)
    ws.cell(row=tot,column=15,value="wtd →").font=Font(name=FONT,italic=True,size=9,color="666666")
    for r in range(first,last+1):
        ws.cell(row=r,column=4).number_format=FMT_NUM0
        ws.cell(row=r,column=5).number_format=FMT_PCT2
        ws.cell(row=r,column=6).number_format='0'
        ws.cell(row=r,column=7).number_format=FMT_PCT2
        ws.cell(row=r,column=8).number_format=FMT_PCT2
        ws.cell(row=r,column=12).number_format='#,##0'
        ws.cell(row=r,column=14).number_format=FMT_PCT2
        ws.cell(row=r,column=15).number_format='0.00'
        ws.cell(row=r,column=16).number_format='0.00'
        ws.cell(row=r,column=17).number_format=FMT_NUM0
        ws.cell(row=r,column=18).number_format=FMT_NUM0
        for cc in (19,20): ws.cell(row=r,column=cc).number_format=FMT_USD
        ws.cell(row=r,column=21).number_format=FMT_PCT
        ws.cell(row=r,column=22).number_format=FMT_PCT2
        ws.cell(row=r,column=23).number_format=FMT_USD
        for cc in (10,11,13): ws.cell(row=r,column=cc).number_format=FMT_DATE
    ws.cell(row=tot,column=16).number_format='0.00'
    ws.cell(row=tot,column=19).number_format=FMT_USD
    ws.cell(row=tot,column=20).number_format=FMT_USD
    ws.cell(row=tot,column=23).number_format=FMT_USD
    col_widths(ws,[16,15,20,11,9,6,8,8,6,12,12,10,12,8,9,8,13,12,15,14,10,8,16])
    return f"'{ws.title}'!$S${tot}", f"'{ws.title}'!$T${tot}", f"'{ws.title}'!$W${tot}"

bi = wb.create_sheet("BondsIntl")
BI_TOT, BI_ACC, BI_PNL = build_bonds(bi, "International Bonds",
    "Accrued = Face x Coupon x days-since-last / 365. YTM/Duration via YIELD/DURATION. Unrealized P&L = Face x (Price - Cost) / 100.", DATA["bonds_intl"])
ba = wb.create_sheet("BondsARG")
BA_TOT, BA_ACC, BA_PNL = build_bonds(ba, "Argentine Bonds",
    "Hard-dollar (GD30/AL30) and CER/ARS bonds. Note: nominal YTM/duration are not meaningful for CER-linked bonds (TX26).", DATA["bonds_arg"])

# ---------- Structured products ----------
st = wb.create_sheet("Structured")
hr = sheet_start(st, "Structured Products",
    "Notes that accrue interest over time and pay coupons every period. Accrual & next-coupon amount recalc live via TODAY().", span=14)
cols=["Product","ISIN","Name","Notional","Accrual % p.a.","Freq/yr","Ccy","Issue","Last Coupon","Next Coupon","Days to Next","Accrued Int.","Next Coupon Amt","Value+Accrued (USD)","Cost %","Mark %","Unreal. P&L (USD)"]
style_header(st, hr, cols)
first=hr+1
for i,s in enumerate(DATA["structured"]):
    r=first+i
    st.cell(row=r,column=1,value=s["id"]).font=C_INPUT
    ic=st.cell(row=r,column=2,value=s.get("isin","")); ic.font=C_INPUT; ic.fill=YELLOW
    st.cell(row=r,column=3,value=s["name"]).font=C_INPUT
    st.cell(row=r,column=4,value=s["notional"]).font=C_INPUT
    ac=st.cell(row=r,column=5,value=s["accrual_pct"]/100); ac.font=C_INPUT
    st.cell(row=r,column=6,value=s["freq_per_yr"]).font=C_INPUT
    st.cell(row=r,column=7,value=s["ccy"]).font=C_INPUT
    st.cell(row=r,column=8,value=s["issue_date"]).font=C_INPUT
    st.cell(row=r,column=9,value=s["last_coupon"]).font=C_INPUT
    st.cell(row=r,column=10,value=s["next_coupon"]).font=C_INPUT
    st.cell(row=r,column=11,value=f"=J{r}-TODAY()").font=C_FORM
    st.cell(row=r,column=12,value=f"=D{r}*E{r}*(TODAY()-I{r})/365").font=C_FORM
    st.cell(row=r,column=13,value=f"=D{r}*E{r}/F{r}").font=C_FORM
    st.cell(row=r,column=14,value=f"=(D{r}*P{r}/100+L{r})/{per_usd(f'G{r}')}").font=C_FORM
    co=st.cell(row=r,column=15,value=s.get("cost_pct",100.0)); co.font=C_INPUT; co.fill=YELLOW
    mk=st.cell(row=r,column=16,value=s.get("mark_pct",100.0)); mk.font=C_INPUT; mk.fill=YELLOW
    st.cell(row=r,column=17,value=f"=D{r}*(P{r}-O{r})/100/{per_usd(f'G{r}')}").font=C_FORM
    last=r
tot=last+1
st.cell(row=tot,column=1,value="TOTAL")
st.cell(row=tot,column=12,value=f"=SUM(L{first}:L{last})")
st.cell(row=tot,column=14,value=f"=SUM(N{first}:N{last})")
st.cell(row=tot,column=17,value=f"=SUM(Q{first}:Q{last})")
totals_row_style(st,tot,17)
for r in range(first,last+1):
    st.cell(row=r,column=4).number_format=FMT_NUM0
    st.cell(row=r,column=5).number_format=FMT_PCT2
    st.cell(row=r,column=6).number_format='0'
    st.cell(row=r,column=11).number_format='#,##0'
    st.cell(row=r,column=12).number_format=FMT_NUM0
    st.cell(row=r,column=13).number_format=FMT_NUM0
    st.cell(row=r,column=14).number_format=FMT_USD
    st.cell(row=r,column=15).number_format='0.00'
    st.cell(row=r,column=16).number_format='0.00'
    st.cell(row=r,column=17).number_format=FMT_USD
    for cc in (8,9,10): st.cell(row=r,column=cc).number_format=FMT_DATE
st.cell(row=tot,column=12).number_format=FMT_NUM0
st.cell(row=tot,column=14).number_format=FMT_USD
st.cell(row=tot,column=17).number_format=FMT_USD
col_widths(st,[22,14,20,12,12,7,6,11,12,12,10,12,13,17,7,7,16])
ST_TOT=f"'Structured'!$N${tot}"; ST_ACC=f"'Structured'!$L${tot}"; ST_PNL=f"'Structured'!$Q${tot}"

# ---------- Options ----------
op = wb.create_sheet("Options")
hr=sheet_start(op,"Options","Contracts negative = short. Market value & P&L in USD. Long options are assets; short options net as liabilities.",span=14)
cols=["Position","Underlying","Und. ISIN","Type","Strike","Expiry","Contracts","Mult","Entry Premium","Price (live)","Ccy","Days to Exp","Market Value (USD)","P&L (USD)"]
style_header(op,hr,cols)
first=hr+1
for i,o in enumerate(DATA["options"]):
    r=first+i
    op.cell(row=r,column=1,value=o["id"]).font=C_INPUT
    op.cell(row=r,column=2,value=o["underlying"]).font=C_INPUT
    ic=op.cell(row=r,column=3,value=o.get("underlying_isin","")); ic.font=C_INPUT; ic.fill=YELLOW
    op.cell(row=r,column=4,value=o["type"]).font=C_INPUT
    op.cell(row=r,column=5,value=o["strike"]).font=C_INPUT
    op.cell(row=r,column=6,value=o["expiry"]).font=C_INPUT
    op.cell(row=r,column=7,value=o["contracts"]).font=C_INPUT
    op.cell(row=r,column=8,value=o["multiplier"]).font=C_INPUT
    op.cell(row=r,column=9,value=o["entry_premium"]).font=C_INPUT
    pc=op.cell(row=r,column=10,value=o["price"]); pc.font=C_INPUT; pc.fill=YELLOW
    op.cell(row=r,column=11,value=o["ccy"]).font=C_INPUT
    op.cell(row=r,column=12,value=f"=F{r}-TODAY()").font=C_FORM
    op.cell(row=r,column=13,value=f"=G{r}*H{r}*J{r}/{per_usd(f'K{r}')}").font=C_FORM
    op.cell(row=r,column=14,value=f"=(J{r}-I{r})*G{r}*H{r}/{per_usd(f'K{r}')}").font=C_FORM
    last=r
tot=last+1
op.cell(row=tot,column=1,value="TOTAL")
op.cell(row=tot,column=13,value=f"=SUM(M{first}:M{last})")
op.cell(row=tot,column=14,value=f"=SUM(N{first}:N{last})")
totals_row_style(op,tot,14)
for r in range(first,last+1):
    for cc in (5,9,10): op.cell(row=r,column=cc).number_format=FMT_NUM
    op.cell(row=r,column=7).number_format='#,##0'
    op.cell(row=r,column=8).number_format='#,##0'
    op.cell(row=r,column=12).number_format='#,##0'
    for cc in (13,14): op.cell(row=r,column=cc).number_format=FMT_USD
    op.cell(row=r,column=6).number_format=FMT_DATE
op.cell(row=tot,column=13).number_format=FMT_USD
op.cell(row=tot,column=14).number_format=FMT_USD
col_widths(op,[18,11,14,7,9,12,10,7,13,12,6,11,16,14])
OP_TOT=f"'Options'!$M${tot}"; OP_PNL=f"'Options'!$N${tot}"

# ---------- Futures ----------
fu=wb.create_sheet("Futures")
hr=sheet_start(fu,"Futures","Marked to market. P&L = (Price-Entry) x Qty x Multiplier. Notional shown for exposure; margin is posted collateral.",span=11)
cols=["Contract","Name","Qty","Entry Price","Price (live)","Mult","Margin","Ccy","Notional (USD)","Unreal. P&L (USD)","Margin (USD)"]
style_header(fu,hr,cols)
first=hr+1
for i,f in enumerate(DATA["futures"]):
    r=first+i
    fu.cell(row=r,column=1,value=f["id"]).font=C_INPUT
    fu.cell(row=r,column=2,value=f["name"]).font=C_INPUT
    fu.cell(row=r,column=3,value=f["qty"]).font=C_INPUT
    fu.cell(row=r,column=4,value=f["entry_price"]).font=C_INPUT
    pc=fu.cell(row=r,column=5,value=f["price"]); pc.font=C_INPUT; pc.fill=YELLOW
    fu.cell(row=r,column=6,value=f["multiplier"]).font=C_INPUT
    fu.cell(row=r,column=7,value=f["margin"]).font=C_INPUT
    fu.cell(row=r,column=8,value=f["ccy"]).font=C_INPUT
    fu.cell(row=r,column=9,value=f"=C{r}*E{r}*F{r}/{per_usd(f'H{r}')}").font=C_FORM
    fu.cell(row=r,column=10,value=f"=(E{r}-D{r})*C{r}*F{r}/{per_usd(f'H{r}')}").font=C_FORM
    fu.cell(row=r,column=11,value=f"=G{r}/{per_usd(f'H{r}')}").font=C_FORM
    last=r
tot=last+1
fu.cell(row=tot,column=1,value="TOTAL")
fu.cell(row=tot,column=10,value=f"=SUM(J{first}:J{last})")
fu.cell(row=tot,column=11,value=f"=SUM(K{first}:K{last})")
totals_row_style(fu,tot,11)
for r in range(first,last+1):
    fu.cell(row=r,column=3).number_format='#,##0'
    for cc in (4,5): fu.cell(row=r,column=cc).number_format=FMT_NUM
    fu.cell(row=r,column=6).number_format='#,##0'
    fu.cell(row=r,column=7).number_format=FMT_NUM0
    for cc in (9,10,11): fu.cell(row=r,column=cc).number_format=FMT_USD
fu.cell(row=tot,column=10).number_format=FMT_USD
fu.cell(row=tot,column=11).number_format=FMT_USD
col_widths(fu,[14,20,8,12,13,7,12,7,16,17,14])
FU_TOT=f"'Futures'!$J${tot}"   # P&L as MTM contribution
FU_MARGIN=f"'Futures'!$K${tot}"

# ---------- Loans ----------
ln=wb.create_sheet("Loans")
hr=sheet_start(ln,"Loans / Bank Overdraft","Liabilities. Accrued interest = Principal x Rate x days-since-start / 365 (simple, Actual/365). Subtracted from net worth.",span=8)
cols=["Facility","Principal","Rate % p.a.","Ccy","Start Date","Accrued Int.","Total Owed","Total Owed (USD)"]
style_header(ln,hr,cols)
first=hr+1
for i,l in enumerate(DATA["loans"]):
    r=first+i
    ln.cell(row=r,column=1,value=l["id"]).font=C_INPUT
    ln.cell(row=r,column=2,value=l["principal"]).font=C_INPUT
    rt=ln.cell(row=r,column=3,value=l["rate_pct"]/100); rt.font=C_INPUT
    ln.cell(row=r,column=4,value=l["ccy"]).font=C_INPUT
    ln.cell(row=r,column=5,value=l["start_date"]).font=C_INPUT
    ln.cell(row=r,column=6,value=f"=B{r}*C{r}*(TODAY()-E{r})/365").font=C_FORM
    ln.cell(row=r,column=7,value=f"=B{r}+F{r}").font=C_FORM
    ln.cell(row=r,column=8,value=f"=G{r}/{per_usd(f'D{r}')}").font=C_FORM
    last=r
tot=last+1
ln.cell(row=tot,column=1,value="TOTAL")
ln.cell(row=tot,column=8,value=f"=SUM(H{first}:H{last})")
totals_row_style(ln,tot,8)
for r in range(first,last+1):
    ln.cell(row=r,column=2).number_format=FMT_NUM0
    ln.cell(row=r,column=3).number_format=FMT_PCT2
    ln.cell(row=r,column=6).number_format=FMT_NUM0
    ln.cell(row=r,column=7).number_format=FMT_NUM0
    ln.cell(row=r,column=8).number_format=FMT_USD
    ln.cell(row=r,column=5).number_format=FMT_DATE
ln.cell(row=tot,column=8).number_format=FMT_USD
col_widths(ln,[22,15,12,7,12,15,16,17])
LN_TOT=f"'Loans'!$H${tot}"

# ---------- Ledger ----------
lg=wb.create_sheet("Ledger")
hr=sheet_start(lg,"Transaction Ledger","Every cash movement. Running Balance is per-currency (cash reconciles here). Qty on Buy/Sell rows feeds the Realized P&L (average-cost) calc.",span=10)
cols=["Date","Type","Asset Class","Instrument","Ccy","Cash Flow","Qty","Cash Flow (USD)","Running Bal (acct·ccy)","Realized P&L (USD)"]
style_header(lg,hr,cols)
first=hr+1
led=sorted(DATA["ledger"], key=lambda x:x["date"])
for i,t in enumerate(led):
    r=first+i
    lg.cell(row=r,column=1,value=t["date"]).font=C_INPUT
    lg.cell(row=r,column=2,value=t["type"]).font=C_INPUT
    lg.cell(row=r,column=3,value=t["asset_class"]).font=C_INPUT
    lg.cell(row=r,column=4,value=t["instrument"]).font=C_INPUT
    lg.cell(row=r,column=5,value=t["ccy"]).font=C_INPUT
    lg.cell(row=r,column=6,value=t["cash_flow"]).font=C_INPUT
    if t.get("qty") is not None:
        lg.cell(row=r,column=7,value=t["qty"]).font=C_INPUT
    lg.cell(row=r,column=8,value=f"=F{r}/{per_usd(f'E{r}')}").font=C_FORM
    lg.cell(row=r,column=9,value=f"=SUMIFS($F${first}:F{r},$E${first}:E{r},E{r},$K${first}:K{r},K{r})").font=C_FORM
    # realized P&L on a Sell (avg cost): proceeds - avg_buy_price * qty_sold ; exclude derivatives
    lg.cell(row=r,column=10,value=(
        f'=IF(AND(B{r}="Sell",C{r}<>"Options",C{r}<>"Futures",G{r}>0),'
        f'(F{r}-(-SUMIFS($F${first}:$F${last},$D${first}:$D${last},D{r},$B${first}:$B${last},"Buy")/'
        f'SUMIFS($G${first}:$G${last},$D${first}:$D${last},D{r},$B${first}:$B${last},"Buy"))*G{r})'
        f'/{per_usd(f"E{r}")},"")')).font=C_FORM
    last=r
for r in range(first,last+1):
    lg.cell(row=r,column=1).number_format=FMT_DATE
    lg.cell(row=r,column=6).number_format=FMT_NUM0
    lg.cell(row=r,column=7).number_format=FMT_NUM0
    lg.cell(row=r,column=8).number_format=FMT_USD
    lg.cell(row=r,column=9).number_format=FMT_NUM0
    lg.cell(row=r,column=10).number_format=FMT_USD
col_widths(lg,[12,15,17,22,6,14,10,14,15,16])
LEDGER_FIRST=first; LEDGER_LAST=last
LEDGER_REALIZED=f"Ledger!$J${first}:$J${last}"

# ---------- Cash (derived from Ledger) ----------
csh=wb.create_sheet("Cash")
hr=sheet_start(csh,"Cash Position","Derived from the Ledger (opening balances + all flows). Single source of truth — edit the Ledger, not here.",span=5)
cols=["Currency","Balance","Balance (USD)","Balance (EUR)","Balance (ARS)"]
style_header(csh,hr,cols)
first=hr+1
cash_ccys=["USD","EUR","ARS"]
for i,cc in enumerate(cash_ccys):
    r=first+i
    csh.cell(row=r,column=1,value=cc).font=C_INPUT
    csh.cell(row=r,column=2,value=f"=SUMIFS(Ledger!$F${LEDGER_FIRST}:$F${LEDGER_LAST},Ledger!$E${LEDGER_FIRST}:$E${LEDGER_LAST},A{r})").font=C_LINK
    csh.cell(row=r,column=3,value=f"=B{r}/{per_usd(f'A{r}')}").font=C_FORM
    csh.cell(row=r,column=4,value=f"=C{r}*{EUR_REF}").font=C_FORM
    csh.cell(row=r,column=5,value=f"=C{r}*{ARS_REF}").font=C_FORM
    last=r
tot=last+1
csh.cell(row=tot,column=1,value="TOTAL")
for cc,fmt in ((3,FMT_USD),(4,FMT_USD),(5,FMT_ARS)):
    csh.cell(row=tot,column=cc,value=f"=SUM({get_column_letter(cc)}{first}:{get_column_letter(cc)}{last})").number_format=fmt
totals_row_style(csh,tot,5)
for r in range(first,last+1):
    csh.cell(row=r,column=2).number_format=FMT_NUM0
    csh.cell(row=r,column=3).number_format=FMT_USD
    csh.cell(row=r,column=4).number_format=FMT_USD
    csh.cell(row=r,column=5).number_format=FMT_ARS
col_widths(csh,[12,18,16,16,18])
CASH_USD_TOT=f"'Cash'!$C${tot}"

# =====================================================================
# MASTER SHEET  (built last, references everything)
# =====================================================================
ms = wb.create_sheet("Master")
wb.move_sheet("Master", -(len(wb.sheetnames)-1))  # move to front
title(ms, "MASTER DASHBOARD", f"Consolidated portfolio summary  •  As of {AS_OF}  •  All values recalc from the Rates tab and live price cells.", span=6)

# Net worth headline block
ms["A4"]="NET WORTH"; ms["A4"].font=Font(name=FONT,bold=True,size=13,color="1F2A44")
ms["A5"]="USD"; ms["A6"]="EUR"; ms["A7"]="ARS"
for r in (5,6,7): ms.cell(row=r,column=1).font=BOLD

# Allocation table
hr=9
alloc_cols=["Asset Class","Value (USD)","Value (EUR)","Value (ARS)","Weight %"]
style_header(ms, hr, alloc_cols)
classes=[
    ("Global Equities", EQG_TOT),
    ("Argentine Equities", EQA_TOT),
    ("Funds", FN_TOT),
    ("International Bonds", BI_TOT),
    ("Argentine Bonds", BA_TOT),
    ("Structured Products", ST_TOT),
    ("Options", OP_TOT),
    ("Futures (MTM P&L)", FU_TOT),
    ("Cash", CASH_USD_TOT),
]
first=hr+1
for i,(nm,ref) in enumerate(classes):
    r=first+i
    ms.cell(row=r,column=1,value=nm).font=BOLD
    ms.cell(row=r,column=2,value=f"={ref}").font=C_LINK
    ms.cell(row=r,column=3,value=f"=B{r}*{EUR_REF}").font=C_FORM
    ms.cell(row=r,column=4,value=f"=B{r}*{ARS_REF}").font=C_FORM
    last=r
assets_tot=last+1
ms.cell(row=assets_tot,column=1,value="GROSS ASSETS").font=BOLD
ms.cell(row=assets_tot,column=2,value=f"=SUM(B{first}:B{last})")
ms.cell(row=assets_tot,column=3,value=f"=SUM(C{first}:C{last})")
ms.cell(row=assets_tot,column=4,value=f"=SUM(D{first}:D{last})")
totals_row_style(ms,assets_tot,5)
# weights
for r in range(first,last+1):
    ms.cell(row=r,column=5,value=f"=IF($B${assets_tot}=0,0,B{r}/$B${assets_tot})").number_format=FMT_PCT
# liabilities
liab=assets_tot+1
ms.cell(row=liab,column=1,value="Loans / Overdraft").font=BOLD
ms.cell(row=liab,column=2,value=f"=-{LN_TOT}").font=C_LINK
ms.cell(row=liab,column=3,value=f"=B{liab}*{EUR_REF}").font=C_FORM
ms.cell(row=liab,column=4,value=f"=B{liab}*{ARS_REF}").font=C_FORM
for c in range(1,6):
    ms.cell(row=liab,column=c).fill=NEG_FILL
nw=liab+1
ms.cell(row=nw,column=1,value="NET WORTH").font=Font(name=FONT,bold=True,size=12)
ms.cell(row=nw,column=2,value=f"=B{assets_tot}+B{liab}")
ms.cell(row=nw,column=3,value=f"=C{assets_tot}+C{liab}")
ms.cell(row=nw,column=4,value=f"=D{assets_tot}+D{liab}")
totals_row_style(ms,nw,5)
ms.cell(row=nw,column=2).font=Font(name=FONT,bold=True,size=12)
# formats for alloc table
for r in range(first,nw+1):
    ms.cell(row=r,column=2).number_format=FMT_USD0
    ms.cell(row=r,column=3).number_format=FMT_USD0
    ms.cell(row=r,column=4).number_format=FMT_ARS

# headline net worth references
ms["B5"]=f"=B{nw}"; ms["B5"].number_format=FMT_USD0; ms["B5"].font=Font(name=FONT,bold=True,size=13,color="0F7B3F")
ms["B6"]=f"=C{nw}"; ms["B6"].number_format=FMT_USD0; ms["B6"].font=Font(name=FONT,bold=True,size=13,color="0F7B3F")
ms["B7"]=f"=D{nw}"; ms["B7"].number_format=FMT_ARS; ms["B7"].font=Font(name=FONT,bold=True,size=13,color="0F7B3F")

# Right-side info blocks: Cash, Accrued interest, FX snapshot
info_col=7
ms.cell(row=9,column=info_col,value="CASH BY CURRENCY").font=BOLD
ms.cell(row=9,column=info_col).fill=SUBHDR_FILL
for i,cc in enumerate(["USD","EUR","ARS"]):
    r=10+i
    ms.cell(row=r,column=info_col,value=cc)
    ms.cell(row=r,column=info_col+1,value=f"=Cash!B{hr+1+i}").font=C_LINK
    ms.cell(row=r,column=info_col+1).number_format=FMT_NUM0

acc_r=14
ms.cell(row=acc_r,column=info_col,value="ACCRUED INTEREST (USD)").font=BOLD
ms.cell(row=acc_r,column=info_col).fill=SUBHDR_FILL
acc_items=[("Intl Bonds",BI_ACC),("ARG Bonds",BA_ACC),("Structured",ST_ACC)]
for i,(nm,ref) in enumerate(acc_items):
    r=acc_r+1+i
    ms.cell(row=r,column=info_col,value=nm)
    ms.cell(row=r,column=info_col+1,value=f"={ref}/1").font=C_LINK  # ST_ACC/BA_ACC are ccy totals; see note
    ms.cell(row=r,column=info_col+1).number_format=FMT_USD0
# NOTE: bond accrued totals (BI_ACC/BA_ACC) are already USD (col N). Structured ST_ACC is ccy (USD notional) -> fine as USD here.
acc_tot=acc_r+1+len(acc_items)
ms.cell(row=acc_tot,column=info_col,value="Total Accrued").font=BOLD
ms.cell(row=acc_tot,column=info_col+1,value=f"=SUM({get_column_letter(info_col+1)}{acc_r+1}:{get_column_letter(info_col+1)}{acc_tot-1})").font=BOLD
ms.cell(row=acc_tot,column=info_col+1).number_format=FMT_USD0

fx_r=acc_tot+2
ms.cell(row=fx_r,column=info_col,value="FX SNAPSHOT").font=BOLD
ms.cell(row=fx_r,column=info_col).fill=SUBHDR_FILL
fx_items=[("Display Ccy","=Rates!B"+str(disp_row)),("EUR/USD",EUR_REF.replace("$","")),("ARS/USD (sel.)",ARS_REF.replace("$",""))]
ms.cell(row=fx_r+1,column=info_col,value="Selected ARS rate")
ms.cell(row=fx_r+1,column=info_col+1,value=f"=Rates!B{sel_row}").font=C_LINK
ms.cell(row=fx_r+2,column=info_col,value="ARS per USD")
ms.cell(row=fx_r+2,column=info_col+1,value=f"={ARS_REF}").font=C_LINK
ms.cell(row=fx_r+2,column=info_col+1).number_format=FMT_ARS
ms.cell(row=fx_r+3,column=info_col,value="EUR per USD")
ms.cell(row=fx_r+3,column=info_col+1,value=f"={EUR_REF}").font=C_LINK
ms.cell(row=fx_r+3,column=info_col+1).number_format='#,##0.00000'
ms.cell(row=fx_r+4,column=info_col,value="Futures margin (USD)")
ms.cell(row=fx_r+4,column=info_col+1,value=f"={FU_MARGIN}").font=C_LINK
ms.cell(row=fx_r+4,column=info_col+1).number_format=FMT_USD0

perf_r=fx_r+6
ms.cell(row=perf_r,column=info_col,value="PERFORMANCE").font=BOLD
ms.cell(row=perf_r,column=info_col).fill=SUBHDR_FILL
ms.cell(row=perf_r+1,column=info_col,value="Portfolio YTD")
ms.cell(row=perf_r+1,column=info_col+1,value="=Analysis!$B$9").font=C_LINK
ms.cell(row=perf_r+1,column=info_col+1).number_format=FMT_PCT
ms.cell(row=perf_r+2,column=info_col,value="Income yield")
ms.cell(row=perf_r+2,column=info_col+1,value="=Analysis!$B$22").font=C_LINK
ms.cell(row=perf_r+2,column=info_col+1).number_format=FMT_PCT

col_widths(ms,[22,18,18,20,10,3,22,16])
ms.sheet_view.showGridLines=False

# =====================================================================
# SNAPSHOTS SHEET  (portfolio net-worth history — connector appends here)
# =====================================================================
def to_date(s):
    y,m,d=map(int,s.split('-')); return datetime.date(y,m,d)

snap=wb.create_sheet("Snapshots")
title(snap,"Net-Worth Snapshots","History of net worth WITH the FX of each date, so period returns in EUR/ARS use the rate that applied then (not today's). The connector appends here automatically.",span=8)
style_header(snap,4,["Date","Net Worth (USD)","Period Return","EUR/USD (that date)","ARS/USD MEP (that date)","Net Worth (EUR, hist.)","Net Worth (ARS, hist.)","S&P 500 (level)"])
SNAP_FIRST=5
snaps=sorted(DATA["snapshots"], key=lambda x:x["date"])
r=SNAP_FIRST
for j,sp in enumerate(snaps):
    fxh=sp.get("fx",{})
    dc=snap.cell(row=r,column=1,value=to_date(sp["date"])); dc.font=C_INPUT; dc.number_format=FMT_DATE
    nc=snap.cell(row=r,column=2,value=sp["net_worth_usd"]); nc.font=C_INPUT; nc.number_format=FMT_USD0
    if j>0:
        snap.cell(row=r,column=3,value=f"=IF(B{r-1}=0,0,B{r}/B{r-1}-1)").number_format=FMT_PCT
    e=snap.cell(row=r,column=4,value=fxh.get("EUR_per_USD",1.0)); e.font=C_INPUT; e.number_format='#,##0.00000'
    a=snap.cell(row=r,column=5,value=(fxh.get("ars_rates",{}) or {}).get("mep",1.0)); a.font=C_INPUT; a.number_format=FMT_ARS
    snap.cell(row=r,column=6,value=f"=B{r}*D{r}").font=C_FORM; snap.cell(row=r,column=6).number_format=FMT_USD0
    snap.cell(row=r,column=7,value=f"=B{r}*E{r}").font=C_FORM; snap.cell(row=r,column=7).number_format=FMT_ARS
    spx=snap.cell(row=r,column=8,value=sp.get("sp500")); spx.font=C_INPUT; spx.number_format='#,##0'
    r+=1
# live "today" row — current FX from the Rates tab
tdc=snap.cell(row=r,column=1,value="=TODAY()"); tdc.font=C_FORM; tdc.number_format=FMT_DATE
tnc=snap.cell(row=r,column=2,value="=Master!$B$5"); tnc.font=C_LINK; tnc.number_format=FMT_USD0
snap.cell(row=r,column=3,value=f"=IF(B{r-1}=0,0,B{r}/B{r-1}-1)").number_format=FMT_PCT
snap.cell(row=r,column=4,value=f"={EUR_REF}").font=C_LINK; snap.cell(row=r,column=4).number_format='#,##0.00000'
snap.cell(row=r,column=5,value=f"={ARS_REF}").font=C_LINK; snap.cell(row=r,column=5).number_format=FMT_ARS
snap.cell(row=r,column=6,value=f"=B{r}*D{r}").font=C_FORM; snap.cell(row=r,column=6).number_format=FMT_USD0
snap.cell(row=r,column=7,value=f"=B{r}*E{r}").font=C_FORM; snap.cell(row=r,column=7).number_format=FMT_ARS
_spxnow=DATA.get("benchmarks",{}).get("sp500_now")
sxc=snap.cell(row=r,column=8,value=_spxnow); sxc.font=C_INPUT; sxc.number_format='#,##0'
SNAP_LAST=r
col_widths(snap,[14,18,14,18,20,18,20,14])
snap.sheet_view.showGridLines=False; snap.freeze_panes="A5"
SNAP_A=f"Snapshots!$A${SNAP_FIRST}:$A${SNAP_LAST}"
SNAP_B=f"Snapshots!$B${SNAP_FIRST}:$B${SNAP_LAST}"

# =====================================================================
# ANALYSIS SHEET  (YTD, custom-range return, income yield)
# =====================================================================
an=wb.create_sheet("Analysis")
title(an,"Performance Analysis","Portfolio return (time-weighted on net-worth snapshots), a custom date-range calculator, and income yield. Per-instrument YTD lives on each asset sheet.",span=5)
def alabel(r,t,bold=False):
    c=an.cell(row=r,column=1,value=t); c.font=BOLD if bold else Font(name=FONT)
def aval(r,formula,fmt,link=True):
    c=an.cell(row=r,column=2,value=formula); c.font=C_LINK if link else C_INPUT; c.number_format=fmt
an.cell(row=6,column=1,value="PORTFOLIO RETURN").font=BOLD; an.cell(row=6,column=1).fill=SUBHDR_FILL
alabel(7,"Year-start value"); aval(7,f"=Snapshots!$B${SNAP_FIRST}",FMT_USD0)
alabel(8,"Current net worth"); aval(8,f"=Snapshots!$B${SNAP_LAST}",FMT_USD0)
alabel(9,"YTD return",True); an.cell(row=9,column=2,value="=IF(B7=0,0,B8/B7-1)").number_format=FMT_PCT; an.cell(row=9,column=2).font=BOLD
# benchmark block (columns D/E)
an.cell(row=6,column=4,value="BENCHMARK").font=BOLD; an.cell(row=6,column=4).fill=SUBHDR_FILL
an.cell(row=7,column=4,value="S&P 500 YTD")
an.cell(row=7,column=5,value=f"=IF(Snapshots!$H${SNAP_FIRST}=0,0,Snapshots!$H${SNAP_LAST}/Snapshots!$H${SNAP_FIRST}-1)").number_format=FMT_PCT
an.cell(row=7,column=5).font=C_LINK
an.cell(row=8,column=4,value="Hurdle (fixed)")
an.cell(row=8,column=5,value=DATA.get("benchmarks",{}).get("hurdle_pct",8)/100).number_format=FMT_PCT
an.cell(row=8,column=5).font=C_INPUT; an.cell(row=8,column=5).fill=YELLOW
an.cell(row=9,column=4,value="Excess vs S&P").font=BOLD
an.cell(row=9,column=5,value="=B9-E7").number_format=FMT_PCT; an.cell(row=9,column=5).font=BOLD

an.cell(row=11,column=1,value="CUSTOM DATE RANGE").font=BOLD; an.cell(row=11,column=1).fill=SUBHDR_FILL
alabel(12,"Start date  (edit)"); sd=an.cell(row=12,column=2,value=to_date(DATA["meta"]["year_start"])); sd.font=C_INPUT; sd.fill=YELLOW; sd.number_format=FMT_DATE
alabel(13,"End date  (edit)"); ed=an.cell(row=13,column=2,value="=TODAY()"); ed.font=C_INPUT; ed.fill=YELLOW; ed.number_format=FMT_DATE
alabel(14,"Value at start"); an.cell(row=14,column=2,value=f"=IFERROR(INDEX({SNAP_B},MATCH(B12,{SNAP_A},1)),B7)").font=C_FORM; an.cell(row=14,column=2).number_format=FMT_USD0
alabel(15,"Value at end"); an.cell(row=15,column=2,value=f"=IFERROR(INDEX({SNAP_B},MATCH(B13,{SNAP_A},1)),B8)").font=C_FORM; an.cell(row=15,column=2).number_format=FMT_USD0
alabel(16,"Range return",True); an.cell(row=16,column=2,value="=IF(B14=0,0,B15/B14-1)").number_format=FMT_PCT; an.cell(row=16,column=2).font=BOLD
alabel(17,"Days in range"); an.cell(row=17,column=2,value="=B13-B12").number_format='#,##0'
alabel(18,"Annualized",True); an.cell(row=18,column=2,value="=IFERROR((1+B16)^(365/B17)-1,0)").number_format=FMT_PCT; an.cell(row=18,column=2).font=BOLD

an.cell(row=20,column=1,value="INCOME").font=BOLD; an.cell(row=20,column=1).fill=SUBHDR_FILL
alabel(21,"Coupons + dividends (USD)")
an.cell(row=21,column=2,value=f'=SUMIFS(Ledger!$H${LEDGER_FIRST}:$H${LEDGER_LAST},Ledger!$B${LEDGER_FIRST}:$B${LEDGER_LAST},"Coupon")+SUMIFS(Ledger!$H${LEDGER_FIRST}:$H${LEDGER_LAST},Ledger!$B${LEDGER_FIRST}:$B${LEDGER_LAST},"Dividend")').font=C_LINK
an.cell(row=21,column=2).number_format=FMT_USD0
alabel(22,"Income yield (on net worth)",True); an.cell(row=22,column=2,value="=IF(B8=0,0,B21/B8)").number_format=FMT_PCT; an.cell(row=22,column=2).font=BOLD

an.cell(row=24,column=1,value="P&L SUMMARY (USD)").font=BOLD; an.cell(row=24,column=1).fill=SUBHDR_FILL
alabel(25,"Unrealized — Equities"); an.cell(row=25,column=2,value=f"={EQG_PNL}+{EQA_PNL}").font=C_LINK; an.cell(row=25,column=2).number_format=FMT_USD0
alabel(26,"Unrealized — Bonds"); an.cell(row=26,column=2,value=f"={BI_PNL}+{BA_PNL}").font=C_LINK; an.cell(row=26,column=2).number_format=FMT_USD0
alabel(27,"Unrealized — Structured"); an.cell(row=27,column=2,value=f"={ST_PNL}").font=C_LINK; an.cell(row=27,column=2).number_format=FMT_USD0
alabel(28,"Unrealized — Options + Futures"); an.cell(row=28,column=2,value=f"={OP_PNL}+{FU_TOT}").font=C_LINK; an.cell(row=28,column=2).number_format=FMT_USD0
alabel(29,"Total unrealized P&L",True); an.cell(row=29,column=2,value="=SUM(B25:B28)").number_format=FMT_USD0; an.cell(row=29,column=2).font=BOLD
alabel(30,"Realized P&L (closed trades)"); an.cell(row=30,column=2,value=f"=SUM({LEDGER_REALIZED})").font=C_LINK; an.cell(row=30,column=2).number_format=FMT_USD0
alabel(31,"Income (coupons + dividends)"); an.cell(row=31,column=2,value="=B21").font=C_FORM; an.cell(row=31,column=2).number_format=FMT_USD0
alabel(32,"TOTAL P&L",True); an.cell(row=32,column=2,value="=B29+B30+B31").number_format=FMT_USD0; an.cell(row=32,column=2).font=Font(name=FONT,bold=True,size=12)

an.cell(row=34,column=1,value="Note: portfolio return uses the net-worth snapshots; realized P&L is average-cost from the Ledger's Buy/Sell rows; unrealized is each sheet's mark-vs-cost total.").font=SUB_FONT

# ---- YTD by account & by asset class (uses snapshot breakdowns; the HTML has the full per-slice time series) ----
_snap0=sorted(DATA["snapshots"], key=lambda x:x["date"])[0]
ba0=_snap0.get("by_account",{}); bc0=_snap0.get("by_class",{})
ar0=36
an.cell(row=ar0,column=1,value="YTD BY ACCOUNT (USD)").font=BOLD; an.cell(row=ar0,column=1).fill=SUBHDR_FILL
for c,t in enumerate(["Account","Year-start","Now","YTD"]):
    hh=an.cell(row=ar0+1,column=1+c,value=t); hh.font=Font(name=FONT,bold=True,size=10)
for i,a in enumerate(DATA["accounts"]):
    rr=ar0+2+i
    an.cell(row=rr,column=1,value=a["id"]).font=Font(name=FONT)
    sc=an.cell(row=rr,column=2,value=round(ba0.get(a["id"],0),2)); sc.font=C_INPUT; sc.number_format=FMT_USD0
    an.cell(row=rr,column=3,value=f"=ByAccount!$K${5+i}").font=C_LINK; an.cell(row=rr,column=3).number_format=FMT_USD0
    an.cell(row=rr,column=4,value=f"=IF(B{rr}=0,0,C{rr}/B{rr}-1)").number_format=FMT_PCT

cr0=ar0+2+len(DATA["accounts"])+1
_cls_map=[("Global Equities","B"),("Argentine Equities","C"),("Funds","D"),("International Bonds","E"),
          ("Argentine Bonds","F"),("Structured Products","G"),("Options","H"),("Futures","I"),("Cash","J")]
_ba_tot=5+len(DATA["accounts"])  # ByAccount TOTAL row
an.cell(row=cr0,column=1,value="YTD BY ASSET CLASS (USD)").font=BOLD; an.cell(row=cr0,column=1).fill=SUBHDR_FILL
for c,t in enumerate(["Asset class","Year-start","Now","YTD"]):
    hh=an.cell(row=cr0+1,column=1+c,value=t); hh.font=Font(name=FONT,bold=True,size=10)
for i,(cn,col) in enumerate(_cls_map):
    rr=cr0+2+i
    an.cell(row=rr,column=1,value=cn).font=Font(name=FONT)
    sc=an.cell(row=rr,column=2,value=round(bc0.get(cn,0),2)); sc.font=C_INPUT; sc.number_format=FMT_USD0
    an.cell(row=rr,column=3,value=f"=ByAccount!${col}${_ba_tot}").font=C_LINK; an.cell(row=rr,column=3).number_format=FMT_USD0
    an.cell(row=rr,column=4,value=f"=IF(B{rr}=0,0,C{rr}/B{rr}-1)").number_format=FMT_PCT
an.cell(row=cr0+2+len(_cls_map)+1,column=1,value="Year-start values are the first snapshot's per-account / per-class breakdown (from the Snapshots history). The dashboard's Analysis tab has the full filterable time series.").font=SUB_FONT

col_widths(an,[32,16,16,16,12])
an.sheet_view.showGridLines=False

# =====================================================================
# ACCOUNT columns (appended so existing formulas don't shift) + BY-ACCOUNT sheet
# =====================================================================
def append_account(ws, lst, col):
    h=ws.cell(row=4, column=col, value="Account"); h.font=HDR_FONT; h.fill=HDR_FILL
    h.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i,item in enumerate(lst):
        c=ws.cell(row=5+i, column=col, value=item.get("account","")); c.font=C_INPUT
    ws.column_dimensions[get_column_letter(col)].width=12

append_account(wb["EqGlobal"],  DATA["equities_global"], 15)   # O
append_account(wb["EqARG"],     DATA["equities_arg"],    15)   # O
append_account(wb["Funds"],     DATA.get("funds",[]),    16)   # P
append_account(wb["BondsIntl"], DATA["bonds_intl"],      24)   # X
append_account(wb["BondsARG"],  DATA["bonds_arg"],       24)   # X
append_account(wb["Structured"],DATA["structured"],      18)   # R
append_account(wb["Options"],   DATA["options"],         15)   # O
append_account(wb["Futures"],   DATA["futures"],         12)   # L
append_account(wb["Loans"],     DATA["loans"],            9)   # I
# Ledger (written in date-sorted order to match the sheet)
_led=sorted(DATA["ledger"], key=lambda x:x["date"])
_lh=wb["Ledger"].cell(row=4, column=11, value="Account"); _lh.font=HDR_FONT; _lh.fill=HDR_FILL
for i,t in enumerate(_led):
    wb["Ledger"].cell(row=5+i, column=11, value=t.get("account","")).font=C_INPUT
wb["Ledger"].column_dimensions["K"].width=12

# extra descriptive columns (sector/country on equities & bonds; spot/iv on options)
def append_cols(ws, lst, specs):
    for off,(hdr,key,_col) in enumerate(specs):
        c=ws.cell(row=4, column=specs[0][2]+off, value=hdr); c.font=HDR_FONT; c.fill=HDR_FILL
        c.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i,item in enumerate(lst):
            cc=ws.cell(row=5+i, column=specs[0][2]+off, value=item.get(key,"")); cc.font=C_INPUT
        ws.column_dimensions[get_column_letter(specs[0][2]+off)].width=14
append_cols(wb["EqGlobal"],  DATA["equities_global"], [("Sector","sector",16),("Country","country",17)])
append_cols(wb["EqARG"],     DATA["equities_arg"],    [("Sector","sector",16),("Country","country",17)])
append_cols(wb["BondsIntl"], DATA["bonds_intl"],      [("Sector","sector",25),("Country","country",26)])
append_cols(wb["BondsARG"],  DATA["bonds_arg"],       [("Sector","sector",25),("Country","country",26)])
append_cols(wb["Options"],   DATA["options"],         [("Spot","spot",16),("Impl. Vol","iv",17)])

def rng(name, extra=0):
    return (5, 4+len(DATA.get(name,[]))+extra)
SHEET_ROWS={"EqGlobal":rng("equities_global"),"EqARG":rng("equities_arg"),"Funds":rng("funds"),
    "BondsIntl":rng("bonds_intl"),"BondsARG":rng("bonds_arg"),"Structured":rng("structured"),
    "Options":rng("options"),"Futures":rng("futures"),"Loans":rng("loans"),"Ledger":rng("ledger")}
CLASS_SRCS=[("Global Eq","EqGlobal","K","O"),("ARG Eq","EqARG","K","O"),("Funds","Funds","L","P"),
    ("Intl Bonds","BondsIntl","S","X"),("ARG Bonds","BondsARG","S","X"),
    ("Structured","Structured","N","R"),("Options","Options","M","O"),("Futures","Futures","J","L")]

bya=wb.create_sheet("ByAccount")
title(bya,"Holdings by Account","Market value (USD) split by bank / broker and asset class. Cash = all ledger flows per account; Loans are liabilities (negative). Net = assets + cash - loans.",span=11)
hdr=["Account"]+[c[0] for c in CLASS_SRCS]+["Cash","Loans","NET (USD)"]
style_header(bya,4,hdr)
first=5
accts=DATA["accounts"]
for i,a in enumerate(accts):
    r=first+i
    bya.cell(row=r,column=1,value=a["id"]).font=BOLD
    for j,(nm,sh,mv,ac) in enumerate(CLASS_SRCS):
        f0,f1=SHEET_ROWS[sh]
        bya.cell(row=r,column=2+j,value=f"=SUMIFS('{sh}'!${mv}${f0}:${mv}${f1},'{sh}'!${ac}${f0}:${ac}${f1},$A{r})").number_format=FMT_USD0
    ncls=len(CLASS_SRCS)                # class cols occupy B .. (1+ncls)
    cash_c=2+ncls; loan_c=3+ncls; net_c=4+ncls
    lf0,lf1=SHEET_ROWS["Ledger"]
    bya.cell(row=r,column=cash_c,value=f"=SUMIFS(Ledger!$H${lf0}:$H${lf1},Ledger!$K${lf0}:$K${lf1},$A{r})").number_format=FMT_USD0
    nf0,nf1=SHEET_ROWS["Loans"]
    bya.cell(row=r,column=loan_c,value=f"=-SUMIFS(Loans!$H${nf0}:$H${nf1},Loans!$I${nf0}:$I${nf1},$A{r})").number_format=FMT_USD0
    bya.cell(row=r,column=net_c,value=f"=SUM(B{r}:{get_column_letter(cash_c)}{r})+{get_column_letter(loan_c)}{r}").number_format=FMT_USD0
    bya.cell(row=r,column=net_c).font=BOLD
    last=r
tot=last+1
ncls=len(CLASS_SRCS); net_c=4+ncls
bya.cell(row=tot,column=1,value="TOTAL")
for c in range(2,net_c+1):
    L=get_column_letter(c)
    bya.cell(row=tot,column=c,value=f"=SUM({L}{first}:{L}{last})").number_format=FMT_USD0
totals_row_style(bya,tot,net_c)
col_widths(bya,[12]+[12]*ncls+[12,12,15])
bya.sheet_view.showGridLines=False
wb.move_sheet("ByAccount", -(len(wb.sheetnames)-2))  # place just after Master

# =====================================================================
# INCOME (forward schedule + maturity ladder) and RISK (exposures) — computed
# =====================================================================
import calendar
_ar=DATA["fx"]["ars_rate_selected"]
def _pu(ccy): return {"USD":1.0,"EUR":DATA["fx"]["EUR_per_USD"],"GBP":DATA["fx"]["GBP_per_USD"],"ARS":DATA["fx"]["ars_rates"][_ar]}.get(ccy,1.0)
def _tu(a,ccy): return a/_pu(ccy)
_today=datetime.date.fromisoformat(AS_OF)
def _dsince(s): return (_today-datetime.date.fromisoformat(s)).days
def _addm(d,n):
    m=d.month-1+n; y=d.year+m//12; m=m%12+1; return datetime.date(y,m,min(d.day,calendar.monthrange(y,m)[1]))

# forward income schedule (12 months)
_sched=[]
def _stream(nxt,mat,freq,amt,ccy,name,acct,typ):
    if not nxt or not freq or not amt: return
    end=_addm(_today,12); dt=datetime.date.fromisoformat(nxt); m=datetime.date.fromisoformat(mat) if mat else end
    step=max(1,round(12/freq)); g=0
    while dt<=end and dt<=m and g<80:
        if dt>=_today: _sched.append((dt.isoformat(),name,acct,typ,_tu(amt,ccy)))
        dt=_addm(dt,step); g+=1
for b in DATA["bonds_intl"]+DATA["bonds_arg"]:
    _stream(b["next_coupon"],b["maturity"],b["freq_per_yr"],b["face"]*(b["coupon_pct"]/100)/b["freq_per_yr"],b["ccy"],b["id"],b.get("account",""),"Coupon")
for s in DATA["structured"]:
    _stream(s["next_coupon"],s["maturity"],s["freq_per_yr"],s["notional"]*(s["accrual_pct"]/100)/s["freq_per_yr"],s["ccy"],s["id"],s.get("account",""),"Note coupon")
for p in DATA["equities_global"]+DATA["equities_arg"]:
    if p.get("annual_div_ps") and p.get("next_div"):
        _stream(p["next_div"],None,p.get("div_freq",4),p["qty"]*p["annual_div_ps"]/p.get("div_freq",4),p["ccy"],p["ticker"],p.get("account",""),"Dividend")
_sched.sort(key=lambda x:x[0])

inc=wb.create_sheet("Income")
title(inc,"Forward Income & Maturities","Projected coupons, note payments and dividends for the next 12 months, plus a maturity ladder. Computed from the data — regenerated whenever the workbook is rebuilt.",span=5)
style_header(inc,4,["Date","Instrument","Account","Type","Amount (USD)"])
r=5
for d,name,acct,typ,amt in _sched:
    inc.cell(row=r,column=1,value=d).number_format=FMT_DATE
    inc.cell(row=r,column=2,value=name); inc.cell(row=r,column=3,value=acct); inc.cell(row=r,column=4,value=typ)
    inc.cell(row=r,column=5,value=round(amt,2)).number_format=FMT_USD0; r+=1
tot=r
inc.cell(row=tot,column=1,value="TOTAL 12-mo").font=BOLD
inc.cell(row=tot,column=5,value=f"=SUM(E5:E{r-1})").number_format=FMT_USD0
totals_row_style(inc,tot,5)
# maturity ladder
lad_r=tot+2
inc.cell(row=lad_r,column=1,value="MATURITY LADDER (principal returning)").font=BOLD; inc.cell(row=lad_r,column=1).fill=SUBHDR_FILL
style_header(inc,lad_r+1,["Year","Instrument","Account","Maturity","Principal (USD)"])
_mat=[]
for b in DATA["bonds_intl"]+DATA["bonds_arg"]:
    _mat.append((b["maturity"][:4],b["id"],b.get("account",""),b["maturity"],_tu(b["face"],b["ccy"])))
for s in DATA["structured"]:
    _mat.append((s["maturity"][:4],s["id"],s.get("account",""),s["maturity"],_tu(s["notional"],s["ccy"])))
_mat.sort(key=lambda x:x[3])
rr=lad_r+2
for yr,name,acct,mat,pr in _mat:
    inc.cell(row=rr,column=1,value=yr); inc.cell(row=rr,column=2,value=name); inc.cell(row=rr,column=3,value=acct)
    inc.cell(row=rr,column=4,value=mat).number_format=FMT_DATE; inc.cell(row=rr,column=5,value=round(pr,2)).number_format=FMT_USD0; rr+=1
col_widths(inc,[14,26,12,12,16]); inc.sheet_view.showGridLines=False; inc.freeze_panes="A5"

# ---- Risk: exposures + concentration ----
def _accrued_bond(b): return b["face"]*(b["coupon_pct"]/100)*_dsince(b["last_coupon"])/365
_hold=[]
for p in DATA["equities_global"]+DATA["equities_arg"]:
    v=_tu(p["qty"]*p["price"],p["ccy"]); _hold.append((p["ticker"],"Equity",p.get("account",""),p["ccy"],p.get("sector","Other"),p.get("country","—"),v,v))
for p in DATA.get("funds",[]):
    v=_tu(p["units"]*p["nav"],p["ccy"]); _hold.append((p["id"],"Fund",p.get("account",""),p["ccy"],p.get("sector","Fund"),p.get("country","—"),v,v))
for b in DATA["bonds_intl"]+DATA["bonds_arg"]:
    v=_tu(b["face"]*b["price_pct"]/100+_accrued_bond(b),b["ccy"]); _hold.append((b["id"],"Bond",b.get("account",""),b["ccy"],b.get("sector","Government"),b.get("country","—"),v,v))
for s in DATA["structured"]:
    v=_tu(s["notional"]*s.get("mark_pct",100)/100+s["notional"]*(s["accrual_pct"]/100)*_dsince(s["last_coupon"])/365,s["ccy"]); _hold.append((s["id"],"Structured",s.get("account",""),s["ccy"],s.get("sector","Structured Credit"),s.get("country","United States"),v,v))
for o in DATA["options"]:
    v=o["contracts"]*o["multiplier"]*o["price"]/_pu(o["ccy"]); _hold.append((o["id"],"Option",o.get("account",""),o["ccy"],"Derivatives",o.get("country","United States"),v,v))
for f in DATA["futures"]:
    pnl=(f["price"]-f["entry_price"])*f["qty"]*f["multiplier"]/_pu(f["ccy"]); notl=f["qty"]*f["price"]*f["multiplier"]/_pu(f["ccy"]); _hold.append((f["id"],"Future",f.get("account",""),f["ccy"],"Derivatives",f.get("country","United States"),pnl,notl))
for cc in ("USD","EUR","ARS"):
    bal=sum(t["cash_flow"] for t in DATA["ledger"] if t["ccy"]==cc)
    if abs(bal)>0.01: v=_tu(bal,cc); _hold.append((cc+" Cash","Cash","—",cc,"Cash",{"USD":"United States","ARS":"Argentina","EUR":"Eurozone"}[cc],v,v))
for l in DATA["loans"]:
    v=-_tu(l["principal"]+l["principal"]*(l["rate_pct"]/100)*_dsince(l["start_date"])/365,l["ccy"]); _hold.append((l["id"],"Loan",l.get("account",""),l["ccy"],"Financing",{"USD":"United States","ARS":"Argentina","EUR":"Eurozone"}.get(l["ccy"],l["ccy"]),v,v))
_nw=sum(h[6] for h in _hold)
def _expo(idx):
    g={}
    for h in _hold: g[h[idx]]=g.get(h[idx],0)+h[7]
    return sorted(g.items(), key=lambda x:-abs(x[1]))

rk=wb.create_sheet("Risk")
title(rk,"Risk & Exposure","Economic exposure by currency, sector and country, plus concentration. The dashboard's Risk tab adds volatility, drawdown, Sharpe, Greeks and leverage. Computed snapshot.",span=5)
rk.cell(row=4,column=1,value=f"Net worth (USD): {_nw:,.0f}").font=BOLD
def _expo_block(startrow, ttl, idx):
    rk.cell(row=startrow,column=1,value=ttl).font=BOLD; rk.cell(row=startrow,column=1).fill=SUBHDR_FILL
    style_header(rk,startrow+1,["Bucket","Exposure (USD)","% of NW"])
    rr=startrow+2
    for k,v in _expo(idx):
        rk.cell(row=rr,column=1,value=k); rk.cell(row=rr,column=2,value=round(v,2)).number_format=FMT_USD0
        rk.cell(row=rr,column=3,value=(v/_nw if _nw else 0)).number_format=FMT_PCT; rr+=1
    return rr+1
nr=_expo_block(6,"CURRENCY EXPOSURE",3)
nr=_expo_block(nr,"SECTOR EXPOSURE",4)
nr=_expo_block(nr,"COUNTRY EXPOSURE",5)
# concentration
rk.cell(row=nr,column=1,value="CONCENTRATION — TOP 10").font=BOLD; rk.cell(row=nr,column=1).fill=SUBHDR_FILL
style_header(rk,nr+1,["Position","Market value (USD)","Weight"])
_long=sorted([h for h in _hold if h[6]>0 and h[1]!="Cash"], key=lambda x:-x[6]); _lt=sum(h[6] for h in _long) or 1
rr=nr+2
for h in _long[:10]:
    rk.cell(row=rr,column=1,value=f"{h[0]} ({h[1]}, {h[2]})"); rk.cell(row=rr,column=2,value=round(h[6],2)).number_format=FMT_USD0
    rk.cell(row=rr,column=3,value=h[6]/_lt).number_format=FMT_PCT; rr+=1
_hhi=sum((h[6]/_lt)**2 for h in _long)
rk.cell(row=rr+1,column=1,value=f"HHI {_hhi*10000:.0f} · effective # of positions ≈ {1/_hhi if _hhi else 0:.1f} · top-5 {sum(h[6] for h in _long[:5])/_lt*100:.0f}%").font=SUB_FONT
col_widths(rk,[34,18,10]); rk.sheet_view.showGridLines=False

# freeze panes on data sheets
for name in ["EqGlobal","EqARG","Funds","BondsIntl","BondsARG","Structured","Options","Futures","Loans","Ledger","Cash"]:
    wb[name].sheet_view.showGridLines=False
    wb[name].freeze_panes="A5"

wb.save(OUT_PATH)
print("saved. sheets:", wb.sheetnames)
