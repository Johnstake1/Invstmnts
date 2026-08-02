#!/usr/bin/env python3
"""
refresh_prices.py  —  Live-price connector + Excel/HTML sync tool
==================================================================
Runs ON YOUR MACHINE (it needs internet + your broker/IBKR login — a static
dashboard can't reach those safely). It:

  1. Pulls FX  (dolarapi.com  + frankfurter.dev)            -> public, no key
  2. Pulls ARG bonds/equities  (data912.com public API)     -> public, no key
       ...optionally via pyhomebroker with your broker login -> needs creds
  3. Pulls global equities/ETFs (Yahoo/Stooq, no key)       -> works out of the box
       ...plus options & futures via IBKR when the Gateway is up (no wiring needed).
  4. Pulls international bond prices via a bond API          -> Finnhub (free key)
       ...falls back to leaving them as manual inputs.
  5. Writes everything back into  dataset.json,
     regenerates  Investment_Dashboard.html,
     updates the yellow price cells in  Investment_Master.xlsx,
     and appends a net-worth row to the Snapshots history.

So one command keeps the Excel engine and the HTML view perfectly in sync.

Quick start
-----------
    pip install requests openpyxl
    # optional, depending on which sources you turn on:
    pip install ib_insync pyhomebroker

    python refresh_prices.py --sources fx,arg_public,intl        # no logins needed
    python refresh_prices.py --sources fx,arg_public,ibkr,intl   # + IBKR gateway
    python refresh_prices.py --no-network                        # just re-sync files

Credentials
-----------
Put all secrets in  credentials.env  (same folder) — it's auto-loaded at startup.
    python refresh_prices.py --check-config        # shows what's configured (no secrets printed)

Alerts & scheduling
-------------------
    python refresh_prices.py --alerts-only        # print/send coupons due, expiries, maturities, big moves, loan interest
Notifications go to every channel filled in credentials.env: email (ALERT_SMTP_*/ALERT_TO),
Telegram (TELEGRAM_*), WhatsApp (TWILIO_*). Thresholds live in dataset.json -> config.alerts.
Schedule it so it runs on its own:
  - macOS/Linux cron:   0 8 * * *  cd /path/to/folder && /usr/bin/python3 refresh_prices.py --alerts-only
  - Windows Task Scheduler: a daily task running the same command
  - Or a Cowork scheduled task that runs the connector and reports back.

Config: edit the CONFIG block below, or set env vars (see each provider).
Everything is best-effort: a source that fails is logged and skipped, the rest
still run. Nothing here is tested in the delivery sandbox (no external network) —
run it locally and adjust the two credentialed providers to your setup.
"""

import os, sys, json, argparse, datetime, urllib.request, urllib.error

HERE = os.path.dirname(os.path.abspath(__file__))
DATASET  = os.path.join(HERE, "dataset.json")
TEMPLATE = os.path.join(HERE, "dashboard_template.html")
HTML_OUT = os.path.join(HERE, "Investment_Dashboard.html")
XLSX     = os.path.join(HERE, "Investment_Master.xlsx")

def load_env():
    """Auto-load KEY=VALUE lines from credentials.env / .env next to this script
    into os.environ. Real environment variables take precedence over the file."""
    for fname in ("credentials.env", ".env"):
        path = os.path.join(HERE, fname)
        if not os.path.exists(path):
            continue
        try:
            with open(path) as fh:
                for line in fh:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    k, v = line.split("=", 1)
                    k, v = k.strip(), v.strip().strip('"').strip("'")
                    if k and v and k not in os.environ:   # don't override real env vars
                        os.environ[k] = v
        except Exception as e:
            print(f"[refresh] ! could not read {fname}: {e}")
load_env()

def _envi(key, default):
    try: return int(os.environ.get(key, str(default)))
    except Exception: return default

# ------------------------------------------------------------------ CONFIG
# All secrets/settings come from credentials.env (or real env vars) — never hardcoded here.
CONFIG = {
    # --- IBKR ---
    "ibkr_mode": os.environ.get("IBKR_MODE", "ib_insync"),   # "ib_insync" | "cpapi"
    "ib_host": os.environ.get("IB_HOST", "127.0.0.1"),
    "ib_port": _envi("IB_PORT", 7497),                        # 7497 paper TWS, 7496 live, 4001/4002 Gateway
    "ib_client_id": _envi("IB_CLIENT_ID", 17),
    "cpapi_base": os.environ.get("CPAPI_BASE", "https://localhost:5000/v1/api"),

    # --- Argentine homebroker (optional; public API used by default) ---
    "hb_broker": _envi("HB_BROKER", 0),                       # pyhomebroker broker id
    "hb_dni":   os.environ.get("HB_DNI", ""),
    "hb_user":  os.environ.get("HB_USER", ""),
    "hb_password": os.environ.get("HB_PASSWORD", ""),

    # --- International bonds ---
    "finnhub_key": os.environ.get("FINNHUB_KEY", ""),         # free key from finnhub.io
}

DATA912 = "https://data912.com"


def log(msg): print(f"[refresh] {msg}", flush=True)

def _get_json(url, timeout=20, headers=None, insecure=False):
    req = urllib.request.Request(url, headers=headers or {"User-Agent": "invest-refresh/1.0"})
    ctx = None
    if insecure:
        import ssl; ctx = ssl.create_default_context(); ctx.check_hostname=False; ctx.verify_mode=ssl.CERT_NONE
    with urllib.request.urlopen(req, timeout=timeout, context=ctx) as r:
        return json.loads(r.read().decode())


# ============================================================= 1. FX
def fetch_fx(data):
    """dolarapi.com (ARS) + frankfurter.dev (EUR/GBP). Public, no key."""
    try:
        arr = _get_json(f"https://dolarapi.com/v1/dolares")
        m = {"oficial":"oficial","blue":"blue","bolsa":"mep","contadoconliqui":"ccl","mayorista":"mayorista"}
        for d in arr:
            if d.get("casa") in m and d.get("venta"):
                data["fx"]["ars_rates"][m[d["casa"]]] = float(d["venta"])
        log(f"FX ARS updated: MEP={data['fx']['ars_rates']['mep']} CCL={data['fx']['ars_rates']['ccl']}")
    except Exception as e:
        log(f"! dolarapi failed: {e}")
    try:
        j = _get_json("https://api.frankfurter.dev/v1/latest?base=USD&symbols=EUR,GBP")
        if j.get("rates", {}).get("EUR"): data["fx"]["EUR_per_USD"] = j["rates"]["EUR"]
        if j.get("rates", {}).get("GBP"): data["fx"]["GBP_per_USD"] = j["rates"]["GBP"]
        log(f"FX EUR/USD updated: {data['fx']['EUR_per_USD']}")
    except Exception as e:
        log(f"! frankfurter failed: {e}")


# ============================================================= 2a. ARG public
def fetch_arg_public(data):
    """data912.com public API — no login. Matches by ticker (symbol)."""
    endpoints = ["/live/arg_bonds", "/live/arg_notes", "/live/arg_corp",
                 "/live/arg_stocks", "/live/arg_cedears"]
    quotes = {}
    for ep in endpoints:
        try:
            for row in _get_json(DATA912 + ep):
                sym = (row.get("symbol") or row.get("ticker") or "").upper()
                px  = row.get("c") or row.get("last") or row.get("close") or row.get("price")
                if sym and px: quotes[sym] = float(px)
        except Exception as e:
            log(f"! data912 {ep} failed: {e}")
    if not quotes:
        return
    # ARG equities (price in ARS)
    for p in data["equities_arg"]:
        if p["ticker"].upper() in quotes:
            p["price"] = quotes[p["ticker"].upper()]
    # ARG bonds — data912 bond prices are per 100 nominal; price_pct is a %
    for b in data["bonds_arg"]:
        key = b["id"].upper()
        if key in quotes:
            b["price_pct"] = quotes[key]          # already ~ price per 100
    log(f"ARG public quotes applied ({len(quotes)} symbols)")


# ============================================================= 2b. ARG homebroker
def fetch_arg_homebroker(data):
    """pyhomebroker — live broker book. Needs your broker id + login.
       pip install pyhomebroker ;  docs: github.com/crapher/pyhomebroker"""
    if not (CONFIG["hb_broker"] and CONFIG["hb_dni"]):
        log("homebroker: no credentials set (HB_BROKER/HB_DNI/HB_USER/HB_PASSWORD) — skipping")
        return
    try:
        from pyhomebroker import HomeBroker
        got = {}
        def on_securities(online, quotes):
            for sym, row in quotes.iterrows():
                try: got[str(sym).upper()] = float(row["last"])
                except Exception: pass
        hb = HomeBroker(CONFIG["hb_broker"], on_securities=on_securities)
        hb.auth.login(dni=CONFIG["hb_dni"], user=CONFIG["hb_user"],
                      password=CONFIG["hb_password"], raise_exception=True)
        hb.online.connect()
        # TODO: subscribe to the boards you need, e.g.:
        # hb.online.subscribe_securities('bluechips', '48hs')
        # hb.online.subscribe_securities('government_bonds', '48hs')
        import time; time.sleep(4)
        hb.online.disconnect()
        for p in data["equities_arg"]:
            if p["ticker"].upper() in got: p["price"] = got[p["ticker"].upper()]
        for b in data["bonds_arg"]:
            if b["id"].upper() in got: b["price_pct"] = got[b["id"].upper()]
        log(f"homebroker quotes applied ({len(got)} symbols)")
    except Exception as e:
        log(f"! homebroker failed: {e}")


# ============================================================= 3. Global instruments
# Turnkey: global equities/ETFs price out of the box via free Yahoo/Stooq quotes
# (no key, no login). If IBKR Gateway/TWS is running it's used first and also
# prices options & futures. Nothing to wire — just add tickers.
def _yahoo_price(ticker):
    try:
        j = _get_json(f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1d&range=1d",
                      headers={"User-Agent": "Mozilla/5.0"})
        return float(j["chart"]["result"][0]["meta"]["regularMarketPrice"])
    except Exception:
        return None

def _stooq_price(ticker):
    try:
        sym = ticker.split(".")[0].lower() + ".us"     # Stooq US suffix
        req = urllib.request.Request(f"https://stooq.com/q/l/?s={sym}&f=sd2t2ohlcv&h&e=csv",
                                     headers={"User-Agent": "invest-refresh/1.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            lines = r.read().decode().strip().splitlines()
        if len(lines) >= 2:
            close = lines[1].split(",")[6]
            return float(close) if close not in ("N/D", "") else None
    except Exception:
        return None

def _free_price(ticker):
    return _yahoo_price(ticker) or _stooq_price(ticker)

def fetch_global(data):
    """Price global equities/ETFs (free Yahoo/Stooq), plus options & futures via IBKR
       when the Gateway is up. Match is by ticker — add a US/global stock and it prices."""
    priced = _ibkr_global(data)                      # tickers IBKR priced (may be empty)
    n = 0
    for p in data["equities_global"]:
        if p["ticker"] in priced:
            continue
        px = _free_price(p["ticker"])
        if px:
            p["price"] = px; n += 1
    if n:
        log(f"global equities: {n} priced via free quotes (Yahoo/Stooq)")
    # keep each option's 'spot' in sync with its live underlying (feeds the Greeks)
    px_by_tkr = {p["ticker"]: p["price"] for p in data["equities_global"]}
    for o in data.get("options", []):
        if o.get("underlying") in px_by_tkr:
            o["spot"] = px_by_tkr[o["underlying"]]
    # funds that carry a market ticker (e.g. an ETF) also price via free quotes
    for f in data.get("funds", []):
        if f.get("ticker"):
            px = _free_price(f["ticker"])
            if px:
                f["nav"] = px

def _ibkr_global(data):
    """Price equities, options and futures via ib_insync. Returns the set of equity
       tickers priced so the free fallback fills the rest. Never raises."""
    done = set()
    if CONFIG["ibkr_mode"] == "cpapi":
        return _ibkr_cpapi(data)
    try:
        from ib_insync import IB, Stock, Option, Future
    except Exception:
        log("IBKR: ib_insync not installed — equities use free quotes (pip install ib_insync to add options/futures)")
        return done
    ib = IB()
    try:
        ib.connect(CONFIG["ib_host"], CONFIG["ib_port"], clientId=CONFIG["ib_client_id"], timeout=8)
    except Exception as e:
        log(f"IBKR: gateway not reachable ({e}) — equities use free quotes")
        return done
    def last_price(contract):
        try:
            if not ib.qualifyContracts(contract):
                return None
            t = ib.reqMktData(contract, "", False, False); ib.sleep(2)
            px = t.last or t.close or t.marketPrice()
            return float(px) if px == px and px else None
        except Exception:
            return None
    for p in data["equities_global"]:                       # match by ticker
        px = last_price(Stock(p["ticker"].split(".")[0], "SMART", p["ccy"]))
        if px:
            p["price"] = px; done.add(p["ticker"])
    for o in data.get("options", []):                       # options by strike/expiry/right
        try:
            px = last_price(Option(o["underlying"], o["expiry"].replace("-", ""),
                                   o["strike"], "C" if o["type"] == "Call" else "P",
                                   "SMART", currency=o["ccy"]))
            if px:
                o["price"] = px
        except Exception as e:
            log(f"! IBKR option {o['id']} failed: {e}")
    for f in data.get("futures", []):                       # futures need ib_symbol/exchange/expiry
        sym, exch, exp = f.get("ib_symbol"), f.get("ib_exchange"), f.get("ib_expiry")
        if not (sym and exch and exp):
            log(f"futures {f['id']}: add ib_symbol/ib_exchange/ib_expiry to auto-price via IBKR")
            continue
        px = last_price(Future(sym, str(exp), exch))
        if px:
            f["price"] = px
    ib.disconnect()
    log(f"IBKR: priced {len(done)} equities + options/futures where available")
    return done

def _ibkr_cpapi(data):
    """IBKR Client Portal Web API (localhost:5000) — resolve conid, then snapshot.
       Returns the set of equity tickers priced. Self-signed cert -> insecure=True."""
    base = CONFIG["cpapi_base"]; done = set()
    try:
        _get_json(f"{base}/iserver/auth/status", insecure=True)
    except Exception as e:
        log(f"IBKR CPAPI: gateway not authenticated ({e}) — equities use free quotes")
        return done
    for p in data["equities_global"]:
        try:
            sr = _get_json(f"{base}/iserver/secdef/search?symbol={p['ticker'].split('.')[0]}", insecure=True)
            conid = (sr or [{}])[0].get("conid")
            if not conid:
                continue
            snap = _get_json(f"{base}/iserver/marketdata/snapshot?conids={conid}&fields=31", insecure=True)
            last = (snap or [{}])[0].get("31")
            if last:
                p["price"] = float(str(last).lstrip("C")); done.add(p["ticker"])
        except Exception:
            continue
    log(f"IBKR CPAPI: priced {len(done)} equities")
    return done


# ============================================================= 4. Intl bonds
def fetch_intl_bonds(data):
    """Finnhub bond price (free key at finnhub.io). Endpoint availability depends
       on your plan; if it 403s, prices stay as manual inputs."""
    key = CONFIG["finnhub_key"]
    if not key:
        log("intl bonds: no FINNHUB_KEY set — leaving prices as manual inputs")
        return
    for b in data["bonds_intl"]:
        isin = b.get("isin", "")
        if not isin: continue
        try:
            j = _get_json(f"https://finnhub.io/api/v1/bond/price?isin={isin}&token={key}")
            px = j.get("c") or (j.get("data") or [{}])[0].get("price")
            if px: b["price_pct"] = float(px)
            log(f"intl bond {b['id']}: {px}")
        except Exception as e:
            log(f"! intl bond {b['id']} ({isin}) failed: {e}")


# ============================================================= net worth (mirror)
def _per_usd(fx, ccy, ars_rate):
    return {"USD":1.0, "EUR":fx["EUR_per_USD"], "GBP":fx["GBP_per_USD"],
            "ARS":fx["ars_rates"][ars_rate]}.get(ccy, 1.0)

def compute_networth(data):
    fx = data["fx"]; ar = fx["ars_rate_selected"]
    tu = lambda amt, ccy: amt / _per_usd(fx, ccy, ar)
    today = datetime.date.today()
    def dsince(s):
        y,m,d = map(int, s.split("-")); return (today - datetime.date(y,m,d)).days
    gross = 0.0
    for p in data["equities_global"] + data["equities_arg"]:
        gross += tu(p["qty"]*p["price"], p["ccy"])
    for f in data.get("funds", []):
        gross += tu(f["units"]*f["nav"], f["ccy"])
    for b in data["bonds_intl"] + data["bonds_arg"]:
        clean = b["face"]*b["price_pct"]/100
        acc = b["face"]*(b["coupon_pct"]/100)*dsince(b["last_coupon"])/365
        gross += tu(clean+acc, b["ccy"])
    for s in data["structured"]:
        acc = s["notional"]*(s["accrual_pct"]/100)*dsince(s["last_coupon"])/365
        gross += tu(s["notional"]*s.get("mark_pct",100)/100 + acc, s["ccy"])
    for o in data["options"]:
        gross += o["contracts"]*o["multiplier"]*o["price"]/_per_usd(fx,o["ccy"],ar)
    for f in data["futures"]:
        gross += (f["price"]-f["entry_price"])*f["qty"]*f["multiplier"]/_per_usd(fx,f["ccy"],ar)
    for cc in ("USD","EUR","ARS"):
        bal = sum(t["cash_flow"] for t in data["ledger"] if t["ccy"]==cc)
        gross += tu(bal, cc)
    liab = 0.0
    for l in data["loans"]:
        acc = l["principal"]*(l["rate_pct"]/100)*dsince(l["start_date"])/365
        liab += tu(l["principal"]+acc, l["ccy"])
    return gross - liab


# ============================================================= sync outputs
def regen_html(data):
    tpl = open(TEMPLATE).read()
    open(HTML_OUT, "w").write(tpl.replace("__DATA__", json.dumps(data)))
    log(f"regenerated {os.path.basename(HTML_OUT)}")

def update_xlsx(data):
    try:
        from openpyxl import load_workbook
    except Exception:
        log("! openpyxl not installed — skipping xlsx update"); return
    if not os.path.exists(XLSX):
        log("! xlsx not found — skipping"); return
    wb = load_workbook(XLSX)
    def set_prices(sheet, key_col, price_col, mapping, pct=False):
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=5):
            k = row[key_col-1].value
            if k in mapping:
                row[price_col-1].value = mapping[k]/100 if pct else mapping[k]
    set_prices("EqGlobal", 1, 6, {p["ticker"]: p["price"] for p in data["equities_global"]})
    set_prices("EqARG",    1, 6, {p["ticker"]: p["price"] for p in data["equities_arg"]})
    set_prices("BondsIntl",1, 7, {b["id"]: b["price_pct"] for b in data["bonds_intl"]}, pct=True)
    set_prices("BondsARG", 1, 7, {b["id"]: b["price_pct"] for b in data["bonds_arg"]}, pct=True)
    set_prices("Options",  1,10, {o["id"]: o["price"] for o in data["options"]})
    set_prices("Futures",  1, 5, {f["id"]: f["price"] for f in data["futures"]})
    # Rates tab: FX
    rs = wb["Rates"]; rs["B5"] = data["fx"]["EUR_per_USD"]
    ars_rows = {"OFICIAL":9,"MAYORISTA":10,"MEP":11,"CCL":12,"BLUE":13}
    for name, r in ars_rows.items():
        v = data["fx"]["ars_rates"].get(name.lower())
        if v: rs.cell(row=r, column=2).value = v
    wb.save(XLSX)
    log(f"updated price + FX cells in {os.path.basename(XLSX)} "
        f"(run the xlsx recalc to refresh formulas)")

def _networth_filtered(data, acct):
    """Net worth (USD) for a single account, or whole portfolio if acct is None."""
    import copy
    sub = copy.deepcopy(data)
    if acct is not None:
        for k in ("equities_global","equities_arg","funds","bonds_intl","bonds_arg","structured","options","futures","loans"):
            sub[k] = [x for x in sub.get(k, []) if x.get("account")==acct]
        sub["ledger"] = [t for t in sub["ledger"] if t.get("account")==acct]
    return compute_networth(sub)

def _networth_by_class(data):
    """Net worth (USD) split by asset class (mirrors the dashboard's classes)."""
    fx=data["fx"]; ar=fx["ars_rate_selected"]
    tu=lambda a,ccy: a/_per_usd(fx,ccy,ar); today=datetime.date.today()
    ds=lambda s:(today-datetime.date(*map(int,s.split("-")))).days
    v={}
    v["Global Equities"]=sum(tu(p["qty"]*p["price"],p["ccy"]) for p in data["equities_global"])
    v["Argentine Equities"]=sum(tu(p["qty"]*p["price"],p["ccy"]) for p in data["equities_arg"])
    v["Funds"]=sum(tu(f["units"]*f["nav"],f["ccy"]) for f in data.get("funds",[]))
    v["International Bonds"]=sum(tu(b["face"]*b["price_pct"]/100+b["face"]*(b["coupon_pct"]/100)*ds(b["last_coupon"])/365,b["ccy"]) for b in data["bonds_intl"])
    v["Argentine Bonds"]=sum(tu(b["face"]*b["price_pct"]/100+b["face"]*(b["coupon_pct"]/100)*ds(b["last_coupon"])/365,b["ccy"]) for b in data["bonds_arg"])
    v["Structured Products"]=sum(tu(s["notional"]*s.get("mark_pct",100)/100+s["notional"]*(s["accrual_pct"]/100)*ds(s["last_coupon"])/365,s["ccy"]) for s in data["structured"])
    v["Options"]=sum(o["contracts"]*o["multiplier"]*o["price"]/_per_usd(fx,o["ccy"],ar) for o in data["options"])
    v["Futures"]=sum((f["price"]-f["entry_price"])*f["qty"]*f["multiplier"]/_per_usd(fx,f["ccy"],ar) for f in data["futures"])
    v["Cash"]=sum(tu(t["cash_flow"],t["ccy"]) for t in data["ledger"])
    v["Loans"]=-sum(tu(l["principal"]+l["principal"]*(l["rate_pct"]/100)*ds(l["start_date"])/365,l["ccy"]) for l in data["loans"])
    return {k:round(x,2) for k,x in v.items()}

def check_alerts(data):
    """Compute portfolio alerts (coupons due, option expiries, maturities, big moves,
       loan interest). Returns a list of (severity, text). Mirrors the dashboard."""
    cfg = data.get("config", {}).get("alerts", {}) or {}
    cdays = cfg.get("coupon_days", 21); edays = cfg.get("expiry_days", 21)
    mdays = cfg.get("maturity_days", 120); movepct = cfg.get("move_pct", 20) / 100.0
    fx = data["fx"]; ar = fx["ars_rate_selected"]
    tu = lambda a, ccy: a / _per_usd(fx, ccy, ar)
    today = datetime.date.today()
    def dd(s):
        y, m, d = map(int, s.split("-")); return (datetime.date(y, m, d) - today).days
    out = []
    # upcoming coupons / note coupons / dividends within cdays
    for b in data["bonds_intl"] + data["bonds_arg"]:
        n = dd(b["next_coupon"])
        if 0 <= n <= cdays:
            out.append(("info", f"Coupon ${tu(b['face']*(b['coupon_pct']/100)/b['freq_per_yr'], b['ccy']):,.0f} from {b['id']} ({b.get('account','')}) in {n}d · {b['next_coupon']}"))
    for s in data["structured"]:
        n = dd(s["next_coupon"])
        if 0 <= n <= cdays:
            out.append(("info", f"Note coupon ${tu(s['notional']*(s['accrual_pct']/100)/s['freq_per_yr'], s['ccy']):,.0f} from {s['id']} in {n}d · {s['next_coupon']}"))
    for p in data["equities_global"] + data["equities_arg"]:
        if p.get("annual_div_ps") and p.get("next_div"):
            n = dd(p["next_div"])
            if 0 <= n <= cdays:
                out.append(("info", f"Dividend ${tu(p['qty']*p['annual_div_ps']/(p.get('div_freq') or 1), p['ccy']):,.0f} from {p['ticker']} in {n}d · {p['next_div']}"))
    # option expiries
    for o in data["options"]:
        n = dd(o["expiry"])
        if 0 <= n <= edays:
            out.append(("warn", f"Option {o['id']} ({o.get('account','')}) expires in {n}d · {o['expiry']}"))
    # bond maturities
    for b in data["bonds_intl"] + data["bonds_arg"]:
        n = dd(b["maturity"])
        if 0 <= n <= mdays:
            out.append(("info", f"{b['id']} matures in {n}d · ${tu(b['face'], b['ccy']):,.0f} principal returns"))
    # big movers (YTD price move)
    for p in data["equities_global"] + data["equities_arg"]:
        soy = p.get("price_soy", p["price"]); y = (p["price"]-soy)/soy if soy else 0
        if abs(y) >= movepct:
            out.append(("crit" if y < 0 else "good", f"{p['ticker']} ({p.get('account','')}) {y*100:+.0f}% YTD"))
    # loan interest
    def dsince(s):
        y, m, d = map(int, s.split("-")); return (today - datetime.date(y, m, d)).days
    for l in data["loans"]:
        acc = l["principal"]*(l["rate_pct"]/100)*dsince(l["start_date"])/365
        if acc > 0:
            out.append(("warn", f"{l['id']} accrued ${tu(acc, l['ccy']):,.0f} interest ({l['rate_pct']}% p.a.)"))
    rank = {"crit": 0, "warn": 1, "good": 2, "info": 3}
    out.sort(key=lambda x: rank.get(x[0], 9))
    return out

def email_alerts(alerts):
    """Optional: email the alert digest. Configure via env:
       ALERT_SMTP_HOST, ALERT_SMTP_PORT, ALERT_SMTP_USER, ALERT_SMTP_PASS, ALERT_TO."""
    host = os.environ.get("ALERT_SMTP_HOST"); to = os.environ.get("ALERT_TO")
    if not (host and to and alerts):
        return False
    import smtplib
    from email.mime.text import MIMEText
    body = "\n".join(f"[{s.upper()}] {t}" for s, t in alerts)
    msg = MIMEText("Portfolio alerts — " + datetime.date.today().isoformat() + "\n\n" + body)
    msg["Subject"] = f"Portfolio alerts ({len(alerts)}) — {datetime.date.today().isoformat()}"
    msg["From"] = os.environ.get("ALERT_SMTP_USER", to); msg["To"] = to
    try:
        with smtplib.SMTP(host, int(os.environ.get("ALERT_SMTP_PORT", "587"))) as srv:
            srv.starttls()
            if os.environ.get("ALERT_SMTP_USER"):
                srv.login(os.environ["ALERT_SMTP_USER"], os.environ.get("ALERT_SMTP_PASS", ""))
            srv.send_message(msg)
        log(f"emailed {len(alerts)} alerts to {to}")
        return True
    except Exception as e:
        log(f"! email failed: {e}"); return False

def _alert_text(alerts):
    return "Portfolio alerts — " + datetime.date.today().isoformat() + "\n\n" + \
           "\n".join(f"[{s.upper()}] {t}" for s, t in alerts)

def send_telegram(alerts):
    """Telegram Bot API. Env: TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID.
       (Create a bot via @BotFather; get chat_id from https://api.telegram.org/bot<TOKEN>/getUpdates.)"""
    tok = os.environ.get("TELEGRAM_BOT_TOKEN"); chat = os.environ.get("TELEGRAM_CHAT_ID")
    if not (tok and chat and alerts):
        return False
    import urllib.parse
    data = urllib.parse.urlencode({"chat_id": chat, "text": _alert_text(alerts)}).encode()
    try:
        urllib.request.urlopen(f"https://api.telegram.org/bot{tok}/sendMessage", data=data, timeout=15)
        log(f"sent {len(alerts)} alerts to Telegram"); return True
    except Exception as e:
        log(f"! telegram failed: {e}"); return False

def send_whatsapp(alerts):
    """WhatsApp via Twilio. Env: TWILIO_SID, TWILIO_TOKEN, TWILIO_WA_FROM (e.g. whatsapp:+14155238886),
       TWILIO_WA_TO (whatsapp:+<your number>). Alternative: Meta WhatsApp Cloud API — see note below."""
    sid = os.environ.get("TWILIO_SID"); tok = os.environ.get("TWILIO_TOKEN")
    frm = os.environ.get("TWILIO_WA_FROM"); to = os.environ.get("TWILIO_WA_TO")
    if not (sid and tok and frm and to and alerts):
        return False
    import urllib.parse, base64
    body = urllib.parse.urlencode({"From": frm, "To": to, "Body": _alert_text(alerts)}).encode()
    req = urllib.request.Request(f"https://api.twilio.com/2010-04-01/Accounts/{sid}/Messages.json", data=body)
    req.add_header("Authorization", "Basic " + base64.b64encode(f"{sid}:{tok}".encode()).decode())
    try:
        urllib.request.urlopen(req, timeout=15)
        log(f"sent {len(alerts)} alerts to WhatsApp"); return True
    except Exception as e:
        log(f"! whatsapp failed: {e}"); return False
    # Meta WhatsApp Cloud API alternative:
    #   POST https://graph.facebook.com/v20.0/<PHONE_NUMBER_ID>/messages
    #   headers: Authorization: Bearer <ACCESS_TOKEN>
    #   json: {"messaging_product":"whatsapp","to":"<to>","type":"text","text":{"body": ...}}

def config_status():
    """Show which notification channels & data sources are configured (no secrets printed)."""
    def ok(*keys): return all(os.environ.get(k) for k in keys)
    rows = [
        ("Email (SMTP)",        ok("ALERT_SMTP_HOST", "ALERT_TO")),
        ("Telegram",            ok("TELEGRAM_BOT_TOKEN", "TELEGRAM_CHAT_ID")),
        ("WhatsApp (Twilio)",   ok("TWILIO_SID", "TWILIO_TOKEN", "TWILIO_WA_FROM", "TWILIO_WA_TO")),
        ("FX (dolarapi + frankfurter)",         True),   # public, no key
        ("ARG bonds/stocks (data912)",          True),   # public, no key
        ("Global equities/ETFs (Yahoo/Stooq)",  True),   # free, no key — turnkey
        ("IBKR options & futures (Gateway)",    ok("IB_PORT") or True),  # used if Gateway is up
        ("Homebroker (ARG login)",  ok("HB_BROKER", "HB_DNI")),
        ("Finnhub (intl bonds)",    ok("FINNHUB_KEY")),
    ]
    cred = os.path.join(HERE, "credentials.env")
    log(f"credentials.env {'found' if os.path.exists(cred) else 'NOT found (using real env vars only)'}")
    for name, on in rows:
        print(f"  [{'ON ' if on else 'off'}] {name}")
    if not any(ok(*k) for k in [("ALERT_SMTP_HOST","ALERT_TO"),("TELEGRAM_BOT_TOKEN","TELEGRAM_CHAT_ID"),
               ("TWILIO_SID","TWILIO_TOKEN","TWILIO_WA_FROM","TWILIO_WA_TO")]):
        log("no notification channel is configured yet — fill in credentials.env")

def notify_alerts(alerts):
    """Dispatch the alert digest to every channel that's configured via env vars."""
    if not alerts:
        return
    sent = []
    if email_alerts(alerts): sent.append("email")
    if send_telegram(alerts): sent.append("telegram")
    if send_whatsapp(alerts): sent.append("whatsapp")
    if not sent:
        log("(no notification channel configured — set ALERT_SMTP_* / TELEGRAM_* / TWILIO_* env vars to push alerts)")

def append_snapshot(data):
    nw = round(compute_networth(data), 2)
    d = datetime.date.today().isoformat()
    by_acct = {a["id"]: round(_networth_filtered(data, a["id"]),2) for a in data.get("accounts",[])}
    by_class = _networth_by_class(data)
    row = {"date": d, "net_worth_usd": nw, "fx": {"EUR_per_USD": data["fx"]["EUR_per_USD"],
           "ars_rates": dict(data["fx"]["ars_rates"])}, "by_account": by_acct, "by_class": by_class}
    spx = data.get("benchmarks", {}).get("sp500_now")   # TODO: refresh from a live index source
    if spx is not None: row["sp500"] = spx
    snaps = data.setdefault("snapshots", [])
    if snaps and snaps[-1]["date"] == d:
        snaps[-1].update(row)
    else:
        snaps.append(row)
    log(f"snapshot {d}: net worth ${nw:,.0f} (per-account + per-class logged)")
    return nw


# ============================================================= main
def main():
    ap = argparse.ArgumentParser(description="Refresh live prices and sync Excel + HTML.")
    ap.add_argument("--sources", default="fx,arg_public,global,intl",
                    help="comma list: fx,arg_public,arg_homebroker,global,intl  (global = IBKR if up + free Yahoo/Stooq fallback)")
    ap.add_argument("--no-network", action="store_true", help="skip all fetches; just re-sync files")
    ap.add_argument("--no-snapshot", action="store_true", help="don't append a net-worth snapshot")
    ap.add_argument("--rebuild-xlsx", action="store_true",
                    help="regenerate the whole workbook via build_xlsx.py (picks up NEW positions added in the dashboard) instead of only refreshing price cells")
    ap.add_argument("--alerts-only", action="store_true",
                    help="just compute & print/email portfolio alerts (no price refresh, no file writes)")
    ap.add_argument("--check-config", action="store_true",
                    help="show which notification channels & data sources are configured (no secrets printed)")
    args = ap.parse_args()

    if args.check_config:
        config_status(); return

    data = json.load(open(DATASET))
    srcs = set(s.strip() for s in args.sources.split(",") if s.strip())

    if args.alerts_only:
        al = check_alerts(data)
        log(f"{len(al)} alert(s):")
        for s, t in al: print(f"  [{s.upper():4}] {t}")
        notify_alerts(al)
        return

    if not args.no_network:
        if "fx" in srcs:             fetch_fx(data)
        if "arg_public" in srcs:     fetch_arg_public(data)
        if "arg_homebroker" in srcs: fetch_arg_homebroker(data)
        if "global" in srcs or "ibkr" in srcs: fetch_global(data)
        if "intl" in srcs:           fetch_intl_bonds(data)
    else:
        log("--no-network: skipping fetches, re-syncing files from current dataset")

    if not args.no_snapshot:
        append_snapshot(data)

    data["meta"]["as_of"] = datetime.date.today().isoformat()
    json.dump(data, open(DATASET, "w"), indent=2, ensure_ascii=False)
    log(f"wrote {os.path.basename(DATASET)}")
    regen_html(data)
    if args.rebuild_xlsx:
        import subprocess, sys as _sys
        bx=os.path.join(HERE,"build_xlsx.py")
        if os.path.exists(bx):
            subprocess.run([_sys.executable, bx], check=False)
            log("rebuilt Investment_Master.xlsx from dataset.json (open in Excel to recalc formulas)")
        else:
            log("! build_xlsx.py not found next to this script — falling back to price-cell update"); update_xlsx(data)
    else:
        update_xlsx(data)
    # alert digest (and email if ALERT_SMTP_* / ALERT_TO env are set)
    al = check_alerts(data)
    if al:
        log(f"{len(al)} alert(s):")
        for s, t in al: print(f"  [{s.upper():4}] {t}")
        notify_alerts(al)
    log("done. Open Investment_Dashboard.html, or recalc the xlsx to refresh its formulas.")

if __name__ == "__main__":
    main()
